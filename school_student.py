"""
school_student.py Tkinter 학교 학생 관리 Ver 1.0_250630
id : admin pw: pine8158!
"""
import os
from tkinter import *
from tkinter import ttk
from PIL import Image,ImageTk
from tkinter import messagebox
import openpyxl

# Color, font
bl_fg='crimson'
lf_bg = 'LightSkyBlue'
rtf_bg = 'DeepSkyBlue'
btn_hlb_bg = 'lightseagreen'
lo_bg='forestgreen'
tit_font = ('Georgia', 36, 'bold')
entry_font = ('맑은 고딕', 12)
btn_font = ('맑은 고딕', 14, 'bold')

# 로그인 창 클래스 추가
class LoginWindow:
    def __init__(self, master):
        self.master = master
        self.master.title("로그인")
        self.master.geometry("300x150+600+300")
        self.master.resizable(False, False)

        self.username = StringVar()
        self.password = StringVar()

        Label(master, text="사용자 ID:").pack(pady=5)
        Entry(master, textvariable=self.username, width=30).pack(pady=2)

        Label(master, text="비밀번호:").pack(pady=5)
        Entry(master, textvariable=self.password, show="*", width=30).pack(pady=2)

        Button(master, text="로그인", command=self.check_login).pack(pady=10)

    def check_login(self):
        # 관리자 ID와 비밀번호를 하드코딩
        if self.username.get() == "admin" and self.password.get() == "pine8158!":
            self.master.destroy() # 로그인 창 닫기
            self.open_main_app() # 메인 애플리케이션 열기
        else:
            messagebox.showerror("로그인 오류", "잘못된 사용자 ID 또는 비밀번호입니다.")

    def open_main_app(self):
        # 메인 학생 관리 시스템 인스턴스 생성 및 실행
        root = Tk()
        style = ttk.Style()
        style.configure("Treeview", font=("맑은 고딕", 10))
        app = student(root)
        root.mainloop()

class student:
    def __init__(self,root):
        # 엑셀 파일 경로 설정
        self.excel_file = './school/StudentDatas.xlsx'
        self.root=root
        self.root.geometry("1530x880+10+10")
        self.root.title("학생 관리 시스템")

        # 변수 정의
        self.var_numb=StringVar()        # 학번
        self.var_name=StringVar()       # 성명
        self.var_gender=StringVar()     # 성별
        self.var_regist=StringVar()     # 등록
        self.var_phone=StringVar()      # 전화번호
        self.var_middle=StringVar()     # 출신중학교
        self.var_jumin1=StringVar()     # 주민번호1
        self.var_jumin2=StringVar()     # 주민번호1
        self.var_address=StringVar()    # 주소
        self.var_post=StringVar()       # 우편번호
        self.var_email=StringVar()      # 이메일
        self.var_fatname=StringVar()    # 부성명
        self.var_fatphone=StringVar()   # 부전화
        self.var_motname=StringVar()    # 모성명
        self.var_motphone=StringVar()   # 모전화
        self.var_club=StringVar()       # 특활부서
        self.var_religion=StringVar()   # 종교
        self.var_hobby=StringVar()      # 특별활동

        # 초기 입력 여부 플래그
        self.is_first_input = True
        self.search_by = StringVar()
        self.search_txt = StringVar()

        # 사진 경로 설정
        self.photo_path='images/'
        self.default_img_path='./images/default_photo.jpg'

        # 위젯 생성
        self.create_widgets()

        # 엑셀 파일 로드 및 데이터 가져오기
        try:
            self.load_excel()
            self.fetch_data()
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 로드 실패: {e}")

    def create_widgets(self):
        # 타이틀 프레임
        lbl_title = Label(self.root, text='STUDENTS MANAGEMENT SYSTEM', font=tit_font, fg=bl_fg, bg=lf_bg)
        lbl_title.place(x=0, y=0, width=1530, height=100)

        # 로고 이미지
        logo=Image.open('images/yushin_logo.jpg')
        logo=logo.resize((80,80),Image.LANCZOS)
        self.photo=ImageTk.PhotoImage(logo)
        self.img_1=Label(self.root,image=self.photo)
        self.img_1.place(x=30,y=5,width=80,height=80)

        # 이미지 로고 프레임
        img_frame = Frame(self.root, relief=RIDGE)
        img_frame.place(x=0, y=100, width=1530, height=120)

        # 첫 번째 이미지
        img_1=Image.open('images/emp5.jpg')
        img_1=img_1.resize((540,120),Image.LANCZOS)
        self.photo1=ImageTk.PhotoImage(img_1)
        self.img_1=Label(img_frame,image=self.photo1)
        self.img_1.place(x=0,y=0,width=540,height=120)

        # 두 번째 이미지
        img_2=Image.open('images/emp2.jpg')
        img_2=img_2.resize((500,160),Image.LANCZOS)
        self.photo2=ImageTk.PhotoImage(img_2)
        self.img_2=Label(img_frame,image=self.photo2)
        self.img_2.place(x=520,y=0,width=520,height=120)

        # 세 번째 이미지
        img_3=Image.open('images/yushin.jpg')
        img_3=img_3.resize((500,120),Image.LANCZOS)
        self.photo3=ImageTk.PhotoImage(img_3)
        self.img_3=Label(img_frame,image=self.photo3)
        self.img_3.place(x=1040,y=0,width=500,height=120)

        # 메인 프레임
        Main_frame = Frame(self.root, relief=RIDGE,bg='white')
        Main_frame.place(x=0, y=220, width=1530, height=580)

        # 학생 정보 상단 프레임
        upper_frame = LabelFrame(Main_frame,bd=0,relief=RIDGE, bg='white', font=('맑은 고딕', 11, 'bold'), fg='black')
        upper_frame.place(x=0, y=10, width=1280, height=300)

        # 학번 라벨 및 입력 필드
        lal_Numb=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="학번:",bg='white')
        lal_Numb.grid(row=0,column=0,padx=5,pady=7,sticky=W)
        txt_Numb=ttk.Entry(upper_frame,textvariable=self.var_numb,width=22,font=('맑은 고딕',12,'bold'))
        txt_Numb.grid(row=0,column=1,padx=5,pady=7)

        # 성명 라벨 및 입력 필드
        lal_Name=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="성명:",bg='white')
        lal_Name.grid(row=0,column=2,padx=5,pady=7,sticky=W)
        txt_Name=ttk.Entry(upper_frame,textvariable=self.var_name,width=22,font=('맑은 고딕',12,'bold'))
        txt_Name.grid(row=0,column=3,padx=5,pady=7)

        # 성별 라벨 및 콤보박스
        lal_Gender=Label(upper_frame,text='성별:',font=('맑은 고딕',12,'bold'),bg='white')
        lal_Gender.grid(row=0,column=4,padx=5,pady=7,sticky=W)
        combo_Gender=ttk.Combobox(upper_frame,textvariable=self.var_gender,font=('맑은 고딕',12,'bold'),width=17,state='readonly')
        combo_Gender['value']=('남','여', '기타')
        combo_Gender.current(0)
        combo_Gender.grid(row=0,column=5,padx=5,pady=7, sticky=W)

        # 재적 라벨 및 콤보박스
        lal_Regist=Label(upper_frame,text='재적:',font=('맑은 고딕',12,'bold'),bg='white')
        lal_Regist.grid(row=0,column=6,padx=5,pady=7,sticky=W)
        combo_Regist=ttk.Combobox(upper_frame,textvariable=self.var_regist,font=('맑은 고딕',12,'bold'),width=17,state='readonly')
        combo_Regist['value']=('재학','전학', '자퇴', '퇴학')
        combo_Regist.current(0)
        combo_Regist.grid(row=0,column=7,padx=5,pady=7, sticky=W)

        # 전화번호 라벨 및 입력 필드
        lal_Phone=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="전화번호:",bg='white')
        lal_Phone.grid(row=1,column=0,padx=5,pady=7,sticky=W)
        txt_Phone=ttk.Entry(upper_frame,textvariable=self.var_phone,width=22,font=('맑은 고딕',12,'bold'))
        txt_Phone.grid(row=1,column=1,sticky=W,padx=2,pady=7)

        # 출신중학교 라벨 및 입력 필드
        lal_Middle=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="출신중학교:",bg='white')
        lal_Middle.grid(row=1,column=2,padx=5,pady=7,sticky=W)
        txt_Middle=ttk.Entry(upper_frame,textvariable=self.var_middle,width=22,font=('맑은 고딕',12,'bold'))
        txt_Middle.grid(row=1,column=3,sticky=W,padx=2,pady=7)

        # 주민번호1 라벨 및 입력 필드
        lal_Jumin1=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="주민번호1:",bg='white')
        lal_Jumin1.grid(row=1,column=4,padx=5,pady=7,sticky=W)
        txt_Jumin1=ttk.Entry(upper_frame,textvariable=self.var_jumin1,width=22,font=('맑은 고딕',12,'bold'))
        txt_Jumin1.grid(row=1,column=5,sticky=W,padx=2,pady=7)

        # 주민번호2 라벨 및 입력 필드
        lal_Jumin2=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="주민번호2:",bg='white')
        lal_Jumin2.grid(row=1,column=6,padx=5,pady=7,sticky=W)
        txt_Jumin2=ttk.Entry(upper_frame,textvariable=self.var_jumin2,width=22,font=('맑은 고딕',12,'bold'))
        txt_Jumin2.grid(row=1,column=7,sticky=W,padx=2,pady=7)

        # 주소 라벨 및 입력 필드
        lal_Address=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="주소:",bg='white')
        lal_Address.grid(row=2,column=0,padx=5,pady=7,sticky=W)
        txt_Address=ttk.Entry(upper_frame,textvariable=self.var_address,width=22,font=('맑은 고딕',12,'bold'))
        txt_Address.grid(row=2,column=1,sticky=W,padx=2,pady=7)

        # 우편번호 라벨 및 입력 필드
        lal_Post=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="우편번호:",bg='white')
        lal_Post.grid(row=2,column=2,padx=5,pady=7,sticky=W)
        txt_Post=ttk.Entry(upper_frame,textvariable=self.var_post,width=22,font=('맑은 고딕',12,'bold'))
        txt_Post.grid(row=2,column=3,sticky=W,padx=2,pady=7)

        # 이메일 라벨 및 입력 필드
        lal_Email=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="이메일:",bg='white')
        lal_Email.grid(row=2,column=4,padx=5,pady=7,sticky=W)
        txt_Email=ttk.Entry(upper_frame,textvariable=self.var_email,width=22,font=('맑은 고딕',12,'bold'))
        txt_Email.grid(row=2,column=5,sticky=W,padx=2,pady=7)

        # 부성명 라벨 및 입력 필드
        lal_Fatname=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="부성명:",bg='white')
        lal_Fatname.grid(row=2,column=6,padx=5,pady=7,sticky=W)
        txt_Fatname=ttk.Entry(upper_frame,textvariable=self.var_fatname,width=22,font=('맑은 고딕',12,'bold'))
        txt_Fatname.grid(row=2,column=7,sticky=W,padx=2,pady=7)

        # 부전화 라벨 및 입력 필드
        lal_Fatphone=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="부전화:",bg='white')
        lal_Fatphone.grid(row=3,column=0,padx=5,pady=7,sticky=W)
        txt_Fatphone=ttk.Entry(upper_frame,textvariable=self.var_fatphone,width=22,font=('맑은 고딕',12,'bold'))
        txt_Fatphone.grid(row=3,column=1,sticky=W,padx=2,pady=7)

        # 모성명 라벨 및 입력 필드
        lal_Motname=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="모성명:",bg='white')
        lal_Motname.grid(row=3,column=2,padx=5,pady=7,sticky=W)
        txt_Motname=ttk.Entry(upper_frame,textvariable=self.var_motname,width=22,font=('맑은 고딕',12,'bold'))
        txt_Motname.grid(row=3,column=3,sticky=W,padx=2,pady=7)

        # 모전화 라벨 및 입력 필드
        lal_Motphone=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="모전화:",bg='white')
        lal_Motphone.grid(row=3,column=4,padx=5,pady=7,sticky=W)
        txt_Motphone=ttk.Entry(upper_frame,textvariable=self.var_motphone,width=22,font=('맑은 고딕',12,'bold'))
        txt_Motphone.grid(row=3,column=5,sticky=W,padx=2,pady=7)

        # 특활부서 라벨 및 입력 필드
        lal_Club=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="특활부서:",bg='white')
        lal_Club.grid(row=3,column=6,padx=5,pady=7,sticky=W)
        txt_Club=ttk.Entry(upper_frame,textvariable=self.var_club,width=22,font=('맑은 고딕',12,'bold'))
        txt_Club.grid(row=3,column=7,sticky=W,padx=2,pady=7)

        # 종교 라벨 및 콤보박스
        lal_Religion=Label(upper_frame,text='종교:',font=('맑은 고딕',12,'bold'),bg='white')
        lal_Religion.grid(row=4,column=0,padx=5,pady=7,sticky=W)
        combo_Religion=ttk.Combobox(upper_frame,textvariable=self.var_religion,font=('맑은 고딕',12,'bold'),width=17,state='readonly')
        combo_Religion['value']=('기독교','불교', '천주교', '무교', '기타')
        combo_Religion.current(0)
        combo_Religion.grid(row=4,column=1,padx=5,pady=7, sticky=W)

        # 취미 라벨 및 입력 필드
        lal_Hooby=Label(upper_frame,font=('맑은 고딕',12,'bold'),text="취미:",bg='white')
        lal_Hooby.grid(row=4,column=2,padx=5,pady=7,sticky=W)
        txt_Hooby=ttk.Entry(upper_frame,textvariable=self.var_hobby,width=22,font=('맑은 고딕',12,'bold'))
        txt_Hooby.grid(row=4,column=3,padx=5,pady=7,sticky=W)

        # 학생 사진 프레임
        self.photo_frame = LabelFrame(Main_frame,bd=0, relief=RIDGE, bg='white')
        self.photo_frame.place(x=1280, y=10, width=250, height=280)

        # 기본 이미지 로드
        self.load_default_image()

        # 버튼 프레임
        button_frame=Frame(upper_frame,bd=0,relief=RIDGE, bg='white')
        button_frame.place(x=0,y=234,width=1530,height=60)

        # 저장 버튼
        btn_add = Button(button_frame, command=self.add_data, text="저장", font=btn_font, width=14, bg=btn_hlb_bg, fg='white')
        btn_add.grid(row=0, column=0, padx=10)

        # 수정 버튼
        btn_update = Button(button_frame, command=self.update_data, text="수정", font=btn_font, width=14, bg=btn_hlb_bg, fg='white')
        btn_update.grid(row=0, column=1, padx=10)

        # 삭제 버튼
        btn_delete = Button(button_frame, command=self.delete_data, text="삭제", font=btn_font, width=14, bg=btn_hlb_bg, fg='white')
        btn_delete.grid(row=0, column=2, padx=10)

        # 리셋 버튼
        btn_clear = Button(button_frame, command=self.clear_data, text="리셋", font=btn_font, width=14, bg=btn_hlb_bg, fg='white')
        btn_clear.grid(row=0, column=3, padx=10)

        # 하단 프레임
        down_frame = LabelFrame(Main_frame,bd=0, bg='red',relief=RIDGE, font=('맑은 고딕', 11, 'bold'), fg='black')
        down_frame.place(x=0, y=300, width=1530, height=470)

        # 검색 프레임
        search_frame = LabelFrame(down_frame,bd=0, relief=RIDGE, bg=rtf_bg, fg='white')
        search_frame.place(x=0, y=0, width=1530, height=60)

        # 검색 콤보박스 (학번, 성명, 전화번호로 검색)
        self.var_com_search = StringVar()
        com_txt_search = ttk.Combobox(search_frame, textvariable=self.var_com_search, state="readonly", font=('맑은 고딕', 14), width=10)
        com_txt_search['value'] = ('학번', '성명','전화번호')
        com_txt_search.current(0)
        com_txt_search.grid(row=0, column=1, padx=5, pady=10, sticky=W)

        # 검색 텍스트 입력
        self.var_search_txt = StringVar()
        txt_search = ttk.Entry(search_frame, textvariable=self.var_search_txt, width=20, font=btn_font)
        txt_search.grid(row=0, column=2, padx=5,pady=10)

        # 검색 버튼
        btn_search = Button(search_frame, text="검색",command=self.search_data, font=btn_font, width=14, bg='green', fg='white')
        btn_search.grid(row=0, column=3, padx=5,pady=8)
        btn_ShowAll = Button(search_frame, text="전부출력",command=self.fetch_data, font=btn_font, width=14, bg='green', fg='white')
        btn_ShowAll.grid(row=0, column=4, padx=5,pady=8)

        # 테이블 프레임
        table_frame=Frame(down_frame,bd=0,relief=RIDGE, bg="white")
        table_frame.place(x=0,y=60,width=1530,height=360)

        # 스크롤바
        scroll_x=ttk.Scrollbar(table_frame,orient=HORIZONTAL)
        scroll_y=ttk.Scrollbar(table_frame,orient=VERTICAL)

        # Treeview (테이블)
        self.student_table=ttk.Treeview(table_frame,columns=('numb','name','gender','regist','phone','middle','jumin1','jumin2','address','post','email','fatname',
            'fatphone','motname','motphone','club','religion','hobby'),xscrollcommand=scroll_x.set,yscrollcommand=scroll_y.set)

        scroll_x.pack(side=BOTTOM,fill=X)
        scroll_y.pack(side=RIGHT,fill=Y)

        scroll_x.config(command=self.student_table.xview)
        scroll_y.config(command=self.student_table.yview)

        # 컬럼 헤더 설정
        self.student_table.heading('numb',text='학번')
        self.student_table.heading('name',text='성명')
        self.student_table.heading('gender',text='성별')
        self.student_table.heading('regist',text='재적')
        self.student_table.heading('phone',text='전화번호')
        self.student_table.heading('middle',text='출신중학교')
        self.student_table.heading('jumin1',text='주민1')
        self.student_table.heading('jumin2',text='주민2')
        self.student_table.heading('address',text='주소')
        self.student_table.heading('post',text='우편번호')
        self.student_table.heading('email',text='이메일')
        self.student_table.heading('fatname',text='부성명')
        self.student_table.heading('fatphone',text='부전화')
        self.student_table.heading('motname',text='모성명')
        self.student_table.heading('motphone',text='모전화')
        self.student_table.heading('club',text='특활부서')
        self.student_table.heading('religion',text='종교')
        self.student_table.heading('hobby',text='취미')

        # 'headings'만 보이도록 설정
        self.student_table['show']='headings'

        # 컬럼 너비 설정
        self.student_table.column('numb',width=60)
        self.student_table.column('name',width=70)
        self.student_table.column('gender',width=50)
        self.student_table.column('regist',width=50)
        self.student_table.column('phone',width=100)
        self.student_table.column('middle',width=100)
        self.student_table.column('jumin1',width=60)
        self.student_table.column('jumin2',width=60)
        self.student_table.column('address',width=150)
        self.student_table.column('post',width=60)
        self.student_table.column('email',width=100)
        self.student_table.column('fatname',width=60)
        self.student_table.column('fatphone',width=100)
        self.student_table.column('motname',width=60)
        self.student_table.column('motphone',width=100)
        self.student_table.column('club',width=100)
        self.student_table.column('religion',width=80)
        self.student_table.column('hobby',width=70)

        self.student_table.pack(fill=BOTH,expand=1)
        # 테이블 행 선택 시 get_cursor 함수 호출
        self.student_table.bind("<ButtonRelease-1>", self.get_cursor)
        # 초기 데이터 추가 (예시)
        self.add_data()

        # 하단 바 프레임
        und_frame=Frame(self.root,bd=0,relief=RIDGE)
        und_frame.place(x=0,y=820,width=1530,height=60)

        # 학교 정보 라벨
        lbl_school=Label(und_frame,font=('맑은 고딕',18,'bold'),text="유신고등학교 Since 1972 Tel : 070-5129-2979 | Fax : 031-211-1614",
            bg = lo_bg, fg = 'white')
        lbl_school.place(x=0, y=0, width=1530, height=60)

    def load_excel(self):
        # 엑셀 파일이 없으면 새로 생성하고 헤더 추가
        if not os.path.exists(self.excel_file):
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.append(["Numb", "Name", "Gender", "Regist", "Phone", "Middle", "Jumin1", "Jumin2", "Address","Post", "Email",
                            "Fatname", "Fatphone", "Motname", "Motphone", "Club", "Religion", "Hobby"])  # 헤더 추가
            self.wb.save(self.excel_file)
        else:
            # 기존 엑셀 파일 로드
            self.wb = openpyxl.load_workbook(self.excel_file)
            self.ws = self.wb.active

        if not self.ws:
            raise Exception("워크시트가 초기화 되지 않음")

    def load_default_image(self):
        # 기본 이미지 로드 및 리사이즈
        self.default_img_path = 'images/default_photo.jpg'
        img_4 = Image.open(self.default_img_path)
        img_4 = img_4.resize((230, 260), Image.LANCZOS)
        self.photo_img = ImageTk.PhotoImage(img_4)

        # 이미지 라벨에 설정
        self.img_4 = Label(self.photo_frame, image=self.photo_img)
        self.img_4.place(x=5, y=10, width=230, height=260)

    def add_data(self):
        # 필수 필드 확인
        if not self.var_numb.get() or not self.var_name.get():
            if not self.is_first_input:
                messagebox.showerror("오류", "모든 필드를 입력해야 합니다.")
            return
        self.is_first_input = False # 첫 입력 플래그 업데이트
        # 현재 입력된 데이터로 행 생성
        row = (self.var_numb.get(),self.var_name.get(),self.var_gender.get(),self.var_regist.get(),self.var_phone.get(),self.var_middle.get(),self.var_jumin1.get(),
            self.var_jumin2.get(),self.var_address.get(),self.var_post.get(),self.var_email.get(),self.var_fatname.get(),self.var_fatphone.get(),self.var_motname.get(),
            self.var_motphone.get(),self.var_club.get(),self.var_religion.get(),self.var_hobby.get())

        # 엑셀에 행 추가 및 저장
        self.ws.append(row)
        self.wb.save(self.excel_file)
        # 데이터 다시 가져와 테이블 업데이트
        self.fetch_data()
        # 입력 필드 초기화
        self.clear_data()

    def update_data(self):
        try:
            # 선택된 항목 확인
            selected_item = self.student_table.selection()

            if not selected_item:
                messagebox.showerror("오류", "업데이트할 레코드를 선택하십시오.")
                return

            selected_item = selected_item[0]
            selected_values = self.student_table.item(selected_item, 'values')

            # 필수 필드 확인
            if not self.var_numb.get() or not self.var_name.get():
                messagebox.showerror("오류", "모든 필드(학번 및 성명)가 필요합니다.")
                return

            # 업데이트될 데이터로 행 생성
            updated_row = (self.var_numb.get(),self.var_name.get(),self.var_gender.get(),self.var_regist.get(),self.var_phone.get(),self.var_middle.get(),self.var_jumin1.get(),
            self.var_jumin2.get(),self.var_address.get(),self.var_post.get(),self.var_email.get(),self.var_fatname.get(),self.var_fatphone.get(),self.var_motname.get(),
            self.var_motphone.get(),self.var_club.get(),self.var_religion.get(),self.var_hobby.get())

            # 엑셀 파일에서 해당 레코드 찾아 업데이트
            for row_idx, excel_row in enumerate(self.ws.iter_rows(min_row=2, values_only=False)):
                if excel_row[0].value == selected_values[0]:  # 학번 기준으로 매칭
                    for col_idx, cell in enumerate(excel_row):
                        cell.value = updated_row[col_idx]  # 셀 값 업데이트
                    break

            # 엑셀 파일 저장
            self.wb.save(self.excel_file)

            # 데이터 다시 가져와 테이블 업데이트
            self.fetch_data()
            # 입력 필드 초기화
            self.clear_data()

            messagebox.showinfo("성공", "레코드가 성공적으로 업데이트되었습니다.")

        except Exception as e:
            messagebox.showerror("오류", f"오류 발생: {e}")

    def load_student_image(self, student_number):
        # 학생 이미지 경로 설정
        image_path = f"images/{student_number}.jpg"

        # 이미지가 존재하면 로드, 없으면 기본 이미지 로드
        if os.path.exists(image_path):
            img = Image.open(image_path)
        else:
            img = Image.open("images/default_photo.jpg")

        # 이미지 리사이즈 및 ImageTk 변환
        img = img.resize((230, 260), Image.LANCZOS)
        self.photo_img = ImageTk.PhotoImage(img)

        # 이미지 라벨 업데이트
        self.img_4.config(image=self.photo_img)
        self.img_4.image = self.photo_img

    def delete_data(self):
        # 선택된 항목의 값 가져오기
        selected_item = self.student_table.selection()[0]
        selected_values = self.student_table.item(selected_item, 'values')

        # 엑셀에서 해당 레코드 찾아 삭제
        for row_idx, excel_row in enumerate(self.ws.iter_rows(min_row=2, values_only=False)):
            if excel_row[0].value == selected_values[0]: # 학번 기준으로 매칭
                self.ws.delete_rows(row_idx + 2) # 행 삭제
                break

        # 엑셀 파일 저장
        self.wb.save(self.excel_file)
        # 데이터 다시 가져와 테이블 업데이트
        self.fetch_data()
        # 입력 필드 초기화
        self.clear_data()

    def search_data(self):
        # 검색 기준 및 검색어 가져오기
        search_by = self.var_com_search.get()
        search_txt = self.var_search_txt.get().strip()

        # 검색 기준에 따른 컬럼 인덱스 설정
        if search_by == '학번':
            column_index = 0
        elif search_by == '성명':
            column_index = 1
        elif search_by == '전화번호':
            column_index = 4
        else:
            messagebox.showerror("오류", "유효한 검색 필드를 선택하십시오!")
            return

        # 테이블 비우기
        self.student_table.delete(*self.student_table.get_children())

        # 엑셀 데이터를 순회하며 검색어와 일치하는 행 찾아 테이블에 삽입
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if search_txt.lower() in str(row[column_index]).lower():
                self.student_table.insert('', END, values=row)

        # 검색 결과가 없을 경우 메시지 표시
        if not self.student_table.get_children():
            messagebox.showinfo("정보", "일치하는 레코드가 없습니다.")

    def fetch_data(self):
        # 테이블 비우기
        self.student_table.delete(*self.student_table.get_children())
        # 엑셀의 모든 데이터 가져와 테이블에 삽입
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            self.student_table.insert('', END, values=row)

    def clear_data(self):
        # 모든 입력 필드 초기화
        self.var_numb.set("")
        self.var_name.set("")
        self.var_gender.set("남") # 기본값 설정
        self.var_regist.set("재학") # 기본값 설정
        self.var_phone.set("")
        self.var_middle.set("")
        self.var_jumin1.set("")
        self.var_jumin2.set("")
        self.var_address.set("")
        self.var_post.set("")
        self.var_email.set("")
        self.var_fatname.set("")
        self.var_fatphone.set("")
        self.var_motname.set("")
        self.var_motphone.set("")
        self.var_club.set("")
        self.var_religion.set("기독교") # 기본값 설정
        self.var_hobby.set("")
        # 기본 이미지로 되돌리기
        self.load_default_image()

    def get_cursor(self, event):
        try:
            # 선택된 행의 데이터 가져오기
            cursor_row = self.student_table.focus()
            contents = self.student_table.item(cursor_row)
            row = contents['values']

            if row:
                # 각 변수에 데이터 설정
                self.var_numb.set(row[0])
                self.var_name.set(row[1])
                self.var_gender.set(row[2])
                self.var_regist.set(row[3])
                self.var_phone.set(row[4])
                self.var_middle.set(row[5])
                self.var_jumin1.set(row[6])
                self.var_jumin2.set(row[7])
                self.var_address.set(row[8])
                self.var_post.set(row[9])
                self.var_email.set(row[10])
                self.var_fatname.set(row[11])
                self.var_fatphone.set(row[12])
                self.var_motname.set(row[13])
                self.var_motphone.set(row[14])
                self.var_club.set(row[15])
                self.var_religion.set(row[16])
                self.var_hobby.set(row[17])

                # 학생 학번 가져오기
                student_number = row[0]

                # 해당 학생 이미지 로드
                self.load_student_image(student_number)
            else:
                messagebox.showerror("오류", "선택된 Data에 자료 없음")
        except IndexError as e:
            messagebox.showerror("오류", f"데이터 접근 에러: {e}")
        except Exception as e:
            messagebox.showerror("오류", f"오류 : {e}")


if __name__=="__main__":
    # 로그인 창을 먼저 띄웁니다.
    login_root = Tk()
    login_app = LoginWindow(login_root)
    login_root.mainloop()