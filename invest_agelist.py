"""
invest_agelist.py   출생 연도별 나이 구하기  Ver 1.0_240516
"""
import  openpyxl as op
from datetime import datetime

def calculate_age(birth_date, current_date):
    birth_year, birth_month, birth_day = map(int, birth_date.split('-'))
    current_year, current_month, current_day = map(int, current_date.split('-'))

    age = current_year - birth_year
    if (current_month, current_day) < (birth_month, birth_day):
        age -= 1

    return age

def find_birth_year(resident_id):
    first_digit = int(resident_id[6])
    if first_digit == 1 or first_digit == 2:
        birth_year = 1900 + int(resident_id[:2])
    elif first_digit == 3 or first_digit == 4:
        birth_year = 2000 + int(resident_id[:2])
    else:
        print("올바른 주민등록번호를 입력하세요.")

    return birth_year


def extract_date(input_date):
    # input_date를 파싱하여 datetime 객체로 변환
    datetime_obj = datetime.strptime(input_date, "%Y-%m-%d %H:%M:%S")

    # 시, 분, 초를 0으로 설정하여 날짜 정보만 남기고 시분초를 제거
    date_only = datetime_obj.replace(hour=0, minute=0, second=0)

    # 제거된 날짜 정보만 문자열로 반환
    return date_only.strftime("%Y-%m-%d")

# 새 워크북을 생성하고 워크시트 가져오기
wb = op.load_workbook(filename = "./school/신상_자료.xlsx", data_only = True)
ws = wb['신상']

wb1 = op.Workbook()
ws1 = wb1.active


tmp = []
for i in range(4, ws.max_row+1):
    tmp.append([])
    for cell in ws[i]:
        tmp[-1].append(cell.value)

ws1["B1"] = "만나이 구하기"
ws1["A2"] = "학번"
ws1["B2"] = "성명"
ws1["C2"] = "생일"
ws1["D2"] = "만나이"

current_s = datetime.today().strftime("%Y-%m-%d")
print("오늘 날짜 : ", current_s)
print("=" * 40)
print("학번  성명     주민번호      생일     나이")
print("-" * 40)
num = 2
for person in tmp:
    num = num + 1
    resident_id =  str(person[6]) + str(person[7])
    print_id =  str(person[6]) +"-"+ str(person[7])
    birth_year = find_birth_year(resident_id)
    birth_date = datetime(birth_year, int(resident_id[2:4]), int(resident_id[4:6]))

    birth_s = birth_date.strftime("%Y-%m-%d")
    age = calculate_age(birth_s, current_s)
    print(person[0], person[1], print_id, birth_s, age)
    ws1.cell(row=num, column=1, value = person[0])
    ws1.cell(row=num, column=2, value = person[1])
    ws1.cell(row=num, column=3, value = birth_s)
    ws1.cell(row=num, column=4, value = age)
# 2행에 헤더 설정 --- (2)
print("=" * 40)
# 파일 저장 --- (6)
print("Data saving....")
wb1.save("./school/invest/출생_연도.xlsx")
