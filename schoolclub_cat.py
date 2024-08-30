"""
schoolclub_cat.py    동아리별 분류  Ver 1.0_240607
"""
# from openpyxl import *
import openpyxl as op
def categoryList(col : int)-> list:
    temp_list = [r[col].value for r in ws]
    del temp_list[0]
    temp_set = set(temp_list)
    name_list = list(temp_set)
    return name_list

def categoryData(col, name):
    total_list=[]
    for r in ws.rows:
        temp_list=[]
        cell_num = len(r)
        if r[col].value == name:
            for n in range(0,cell_num):
                temp_list.append(r[n].value)
        if temp_list != []:
            total_list.append(temp_list)
    return total_list

def writeExcel(name, total_list):
    sht = wb.create_sheet(name)
    i=1
    for data in total_list:
        data_length = len(data)
        for n in range(0, data_length):
            sht.cell(row=i, column=n+1).value = data[n]
        i=i+1


path = "./school/club/취합_동아리.xlsx"
# wb = load_workbook(path)
wb = op.load_workbook(path)
ws = wb.active
pb = wb.sheetnames

col = 2 # 기준 열번호
category = categoryList(col)
category.sort()
tit_list = [['학번', '이름', '동아리반']]  #
for name in category:
    tot_list = categoryData(col, name)     # total-> tot
    total_list = tit_list + tot_list
    writeExcel(name, total_list)
    print(name)

ws.title = '전체'
wb.save("./school/club/동아리_분류.xlsx")

