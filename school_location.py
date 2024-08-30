"""
invest_lacation.py 학생 주소 위치 위도 경도  Ver 1.0_240720
"""
import requests
import openpyxl as op


def excel_to_list(file_path):
    workbook = op.load_workbook(file_path)
    sheet = workbook.active
    excel_data = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        excel_data.append(row[8])  # 8번째 9-1

    return excel_data


def getLatLng(address):
    try:
        url = 'https://dapi.kakao.com/v2/local/search/address.json?query=' + address
        rest_api_key = Your_REST_API_KEY
        header = {'Authorization': 'KakaoAK ' + rest_api_key}
        r = requests.get(url, headers=header)

        if r.status_code == 200:
            result_address = r.json()["documents"][0]["address"]
            return result_address["y"], result_address["x"]
        else:
            print(f"API 오류: {r.status_code}")
            return None
    except Exception as e:
        print(f"주소 처리 중 오류 발생: {str(e)}")
        return None


def getKakaoMapHtml(address_latlng):
    if address_latlng is None:
        return ""

    javascript_key = "1b88911358b247b33e269509d157b7ec"

    result = ""
    result = result + "<div id='map' style='width:300px;height:200px;display:inline-block;'></div>" + "\n"
    result = result + "<script type='text/javascript' src='//dapi.kakao.com/v2/maps/sdk.js?appkey=" + javascript_key + "'></script>" + "\n"
    result = result + "<script>" + "\n"
    result = result + "    var container = document.getElementById('map'); " + "\n"
    result = result + "    var options = {" + "\n"
    result = result + "           center: new kakao.maps.LatLng(" + address_latlng[0] + ", " + address_latlng[1] + ")," + "\n"
    result = result + "           level: 3" + "\n"
    result = result + "    }; " + "\n"
    result = result + "    var map = new kakao.maps.Map(container, options); " + "\n"

    result = result + "    var markerPosition  = new kakao.maps.LatLng(" + address_latlng[0] + ", " + address_latlng[1] + ");  " + "\n"
    result = result + "    var marker = new kakao.maps.Marker({position: markerPosition}); " + "\n"
    result = result + "    marker.setMap(map); " + "\n"

    result = result + "</script>" + "\n"

    return result


# main()
if __name__ == "__main__":
    file_path = "./school/신상_자료.xlsx"
    data_list = excel_to_list(file_path)
    wb = op.load_workbook(file_path)
    ws = wb.active
    cnt = 0
    num = 1
    for address in data_list:
        num = num + 1
        print(f"{num - 1}: {address}")

        address_latlng = getLatLng(address)
        if address_latlng:
            print(f"위도: {address_latlng[0]}, 경도: {address_latlng[1]}")
            ws.cell(row=num+1, column=42, value=float(address_latlng[0]))
            ws.cell(row=num+1, column=43, value=float(address_latlng[1]))

            map_html = getKakaoMapHtml(address_latlng)
        else:
            cnt=cnt+1
            print(f"주소 '{address}'를 처리할 수 없습니다.")
            ws.cell(row=num+1, column=42, value="오류")
            ws.cell(row=num+1, column=43, value="오류")

    wb.save('./school/신상_자료.xlsx')
    print(cnt)
    print("처리 완료. 결과가 Excel 파일에 저장되었습니다.")
