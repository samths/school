"""
school_hwp.py  학교 업무 문서 만들기  Ver 1.0_240816
"""
import sys
import os
import win32com.client as win32  # pip install pywin32
import openpyxl as op
from datetime import datetime
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox, QMessageBox
from PyQt5.QtGui import QFont

def extract_birthdate_and_gender(jumin_no):
    year = jumin_no[:2]
    month = jumin_no[2:4]
    day = jumin_no[4:6]
    gender_code = jumin_no[6]

    if gender_code in ['1', '2']:
        year = "19" + year
    elif gender_code in ['3', '4']:
        year = "20" + year
    else:
        return "유효하지 않은 주민등록번호입니다."

    if gender_code in ['1', '3']:
        gender = "남"
    elif gender_code in ['2', '4']:
        gender = "여"
    else:
        return "유효하지 않은 성별 코드입니다."

    birthdate = f"{year}.{month}.{day}"
    return birthdate, gender

class StudentDocumentApp(QWidget):
    def __init__(self):
        super().__init__()
        self.file_names = []
        self.initUI()

    def initUI(self):
        self.setWindowTitle('학생 업무 문서 만들기')
        self.setGeometry(300, 300, 400, 150)

        font = QFont('맑은 고딕', 10)
        QApplication.setFont(font)

        layout = QVBoxLayout()

        # Document type selection
        doc_layout = QHBoxLayout()
        doc_layout.addWidget(QLabel('문서 유형:'))
        self.doc_combo = QComboBox()
        doc_layout.addWidget(self.doc_combo)
        layout.addLayout(doc_layout)

        # Student ID input and Create document button in the same row
        id_btn_layout = QHBoxLayout()
        id_btn_layout.addWidget(QLabel('학번(담임 ex) 3100):'))
        self.id_input = QLineEdit()
        id_btn_layout.addWidget(self.id_input)
        self.create_btn = QPushButton('문서 생성')
        self.create_btn.clicked.connect(self.create_document)
        id_btn_layout.addWidget(self.create_btn)
        layout.addLayout(id_btn_layout)

        self.setLayout(layout)

        # Load HWP list from Excel
        self.load_hwp_list("./school/hwp_list.xlsx")

    def load_hwp_list(self, file_name):
        try:
            wb = op.load_workbook(file_name)
            st = wb.active
            for row in st.iter_rows(min_row=2, values_only=True):
                self.doc_combo.addItem(row[0])
                self.file_names.append(row[1])
        except Exception as e:
            QMessageBox.critical(self, '오류', f'문서 목록을 불러오는 중 오류가 발생했습니다: {str(e)}')

    def create_document(self):
        try:
            student_id = int(self.id_input.text())
        except ValueError:
            QMessageBox.warning(self, '오류', '학번은 숫자로 입력해야 합니다.')
            return

        result = self.search_by_student_id("./school/신상_자료.xlsx", student_id)
        if result is None:
            QMessageBox.warning(self, '오류', '해당 학번의 학생 정보를 찾을 수 없습니다.')
            return

        file_root = os.path.abspath(os.path.join(os.path.dirname(__file__)))

        prn_no = self.doc_combo.currentIndex()
        file_name = self.file_names[prn_no] + ".hwp"
        file_path = os.path.join(file_root, "school", file_name)

        hwp = win32.gencache.EnsureDispatch("hwpframe.hwpobject")
        hwp.Open(file_path, "HWP", "forceopen:true")

        now = datetime.now()
        year, month, day = now.year, now.month, now.day
        six_day = now.strftime('%y%m%d')

        stnum = str(student_id)
        st_grade = int(stnum[0:1])
        st_class = int(stnum[1:2])
        st_number = int(stnum[2:])
        jumin_no =  str(result[6]) + str(result[7])
        print_no =  str(result[6]) + "-" + str(result[7])
        birthdate, gender = extract_birthdate_and_gender(jumin_no)

        hwp.PutFieldText("연도", str(year))
        hwp.PutFieldText("월", str(month))
        hwp.PutFieldText("일", str(day))
        hwp.PutFieldText("학년", str(st_grade))
        hwp.PutFieldText("반", str(st_class))
        hwp.PutFieldText("번호", str(st_number))
        hwp.PutFieldText("성별", gender)
        hwp.PutFieldText("학번", result[0])
        hwp.PutFieldText("성명", result[1])
        hwp.PutFieldText("생일", birthdate)
        hwp.PutFieldText("전화번호", result[4])
        hwp.PutFieldText("출신중학교", result[5])
        hwp.PutFieldText("주민번호", print_no)
        hwp.PutFieldText("주소", result[8])
        hwp.PutFieldText("우편번호", result[9])
        hwp.PutFieldText("이메일", result[9])
        hwp.PutFieldText("부성명", result[11])
        hwp.PutFieldText("부전화", result[12])
        hwp.PutFieldText("모성명", result[13])
        hwp.PutFieldText("모전화", result[14])
        hwp.PutFieldText("특활부서", result[15])
        hwp.PutFieldText("종교", result[16])
        hwp.PutFieldText("취미", result[17])

        new_file_path = os.path.join(file_root, "./school/hwpdoc/" + self.file_names[prn_no] + "_" + stnum + "_" + result[1] + "_" + six_day + ".hwp")
        hwp.SaveAs(new_file_path)
        hwp.Quit()

        QMessageBox.information(self, '성공', f'문서가 성공적으로 저장되었습니다: {new_file_path}')

    def search_by_student_id(self, file_name, student_id):
        wb = op.load_workbook(file_name)
        st = wb.active
        for row in st.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(student_id):
                return row
        return None

if __name__ == '__main__':
    app = QApplication(sys.argv)

    # Set application-wide font
    font = QFont('맑은 고딕', 10)
    app.setFont(font)

    ex = StudentDocumentApp()
    ex.show()
    sys.exit(app.exec_())
