"""
schoolclub_file.py    동아리별 분류 파일  Ver 1.0_240607
분류 파일별 저장
"""
import openpyxl as op
import time
start_time = time.time()

def create_individual_excel_files(input_file):
    # 엑셀 파일 열기
    wb = op.load_workbook(input_file)
    global count_list
    count_list = []
    for sheet_name in wb.sheetnames:
        # 새로운 엑셀 파일 생성
        new_wb =  op.Workbook()
        new_sheet = new_wb.active
        count = -1

        # 원본 시트에서 데이터 복사
        original_sheet = wb[sheet_name]
        for row in original_sheet.iter_rows(values_only=True):
            new_sheet.append(row)
            count += 1

        # 동아리 파일 저장
        output_file = f"{sheet_name}.xlsx"
        new_sheet.title = sheet_name
        new_wb.save("./school/club/동아리_"+output_file)
        print(f"엑셀 파일  '{output_file}' ", end = " ")
        print("의 자료", count,"개가 저장되었습니다.")
        count_r =  [sheet_name, count]
        count_list.append(count_r)

    print("-" * 80)

if __name__ == "__main__":
    input_file = "./school/club/동아리_분류.xlsx"     #  입력 파일명
    create_individual_excel_files(input_file)

    for nam, cnt in count_list:
        print(nam, cnt)

    #  분류 인원 파일 생성
    cnt_wb = op.Workbook()
    cnt_st = cnt_wb.active
    for data in count_list:
        cnt_st.append(data)
    cnt_wb.save("./school/club/분류_인원.xlsx")

end_time = time.time()
print("=" * 30)
print("경과시간 : ", str(end_time-start_time), "seconds")
