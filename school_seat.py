"""
school_seat.py    자리 배정 결과   Ver 1.1_240527
"""
import openpyxl
import random
import pandas


def read_student_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    student_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_id, name = row
        student_data.append({'학번': student_id, '성명': name})
    return student_data

def assign_seats(student_data, seats):
    if len(seats) < len(student_data):
        raise ValueError("자리가 학생 수보다 적습니다. 자리를 더 추가하세요.")

    seat_assignments = {}
    for student in student_data:
        student_id = student['학번']
        assigned_seat = random.choice(seats)
        seat_assignments[student_id] = assigned_seat
        seats.remove(assigned_seat)

    assigned_seats = [{'학번': student['학번'], '성명': student['성명'], '자리': seat_assignments[student['학번']]} for student in student_data]
    return assigned_seats

def print_seat_assignments(assigned_seats):
    for seats in assigned_seats:
        for seat in seats.values():
            print(seat, end=" ")
        print()

    # for seats in assigned_seats:
    #     for subject, score in seats.items():
    #         print(subject, score, end=" ")
    #     print()

def save_to_excel(assigned_seats, output_file_path):
    workbook = openpyxl.Workbook()
    seat_assignment_sheet = workbook.active
    seat_assignment_sheet.title = '자리배정결과'

    # Write headers
    seat_assignment_sheet.append(['', '자리배정', ''])
    seat_assignment_sheet.append([])
    seat_assignment_sheet.append(['학번', '성명', '자리'])

    # Write data
    for assignment in assigned_seats:
        seat_assignment_sheet.append([assignment['학번'], assignment['성명'], assignment['자리']])

    # Create seat location sheet
    seat_location_sheet = workbook.create_sheet(title='자리위치표시')
    # for row in range(2, 7):   # 2.7 1,6
    #     seat_location_sheet.cell(row=row-1, column=1, value=f'열{row-1}')

    # Fill seat locations
    seat_location_sheet.cell(row=1, column=3, value="[교 탁]")
    for assignment in assigned_seats:
        seat = assignment['자리']
        name = assignment['성명']
        row_num = int(seat[1])
        col_num = ord(seat[0]) - ord('A') + 1
        seat_location_sheet.cell(row=row_num+2, column=col_num, value=name)

    workbook.save(output_file_path)

# 파일 경로 설정
file_path = './school/학생데이터.xlsx'
output_file_path = './school/자리배정결과.xlsx'

# 학생 데이터 읽기
student_data = read_student_data(file_path)

# 자리 리스트 준비
seats = ["A1", "A2", "A3", "A4", "A5",
         "B1", "B2", "B3", "B4", "B5",
         "C1", "C2", "C3", "C4", "C5",
         "D1", "D2", "D3", "D4",
         "E1", "E2", "E3", "E4",
         "F1", "F2", "F3", "F4"]

# 자리를 배정하고 결과를 리스트로 받기
assigned_seats = assign_seats(student_data, seats)

# 자리 배정 결과 출력
print_seat_assignments(assigned_seats)

# 결과를 엑셀 파일로 저장
save_to_excel(assigned_seats, output_file_path)

print(f"\n자리 배정 결과가 {output_file_path}에 저장되었습니다.")
