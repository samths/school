"""
school_time.py 학생 등교 시간 기록 Ver 1.0_240823
"""
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QHeaderView
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

today = str(datetime.now().strftime('%Y%m%d'))
stu_now = today[2::]

class ExcelApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # 학생_등교시간.xlsx 파일 경로 설정
        self.excel_path = f"./school/학생_등교시간({stu_now}).xlsx"

        # 학생_등교시간.xlsx 파일 존재 여부 확인
        if not os.path.exists(self.excel_path):
            # 파일이 없으면 신상_자료.xlsx에서 이름 불러오기 및 학생_등교시간.xlsx 생성
            self.create_student_list()

        # 학생_등교시간.xlsx 파일 불러오기
        self.wb = load_workbook(self.excel_path)
        self.sheet = self.wb.active

        # UI 설정
        self.setWindowTitle("학생 출석 기록")
        self.setGeometry(300, 100, 350, 600)

        # 테이블 위젯 생성
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(2)
        self.table_widget.setHorizontalHeaderLabels(["학생 이름", "출석 시간"])

        # 열 폭 조정
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Interactive)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.resizeSection(0, 80)  # 학생 이름 열의 초기 폭을 150픽셀로 설정

        self.table_widget.setRowCount(self.sheet.max_row - 1)

        # 엑셀 파일에서 학생 이름과 출석 시간을 테이블에 추가
        for i, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True)):
            name, attendance_time = row
            self.table_widget.setItem(i, 0, QTableWidgetItem(name))
            self.table_widget.setItem(i, 1, QTableWidgetItem(attendance_time or ""))

        # 테이블 클릭 이벤트 연결
        self.table_widget.cellClicked.connect(self.record_time)

        # 레이아웃 설정
        layout = QVBoxLayout()
        layout.addWidget(self.table_widget)
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def create_student_list(self):
        # 신상_자료.xlsx 파일 불러오기
        source_wb = load_workbook("./school/신상_자료.xlsx")
        source_sheet = source_wb.active

        # 새로운 Workbook 생성
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "학생출석"

        # 헤더 추가
        new_sheet['A1'] = "이름"
        new_sheet['B1'] = "출석 시간"

        # B열에서 이름 복사
        for row in source_sheet.iter_rows(min_row=4, min_col=2, max_col=2, values_only=True):
            new_sheet.append([row[0], None])

        # 학생_출석시간.xlsx 저장
        new_wb.save(self.excel_path)

    def record_time(self, row, column):
        # 학생 이름이 클릭된 경우에만 동작
        if column == 0:
            # 클릭된 학생 이름 가져오기
            name_item = self.table_widget.item(row, column)
            if name_item is not None:
                student_name = name_item.text()

                # 현재 시간 기록
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"{student_name} 출석 시간: {current_time}")

                # 엑셀 파일에서 해당 학생의 출석 시간 기록 및 테이블에 업데이트
                for excel_row in range(2, self.sheet.max_row + 1):
                    if self.sheet.cell(excel_row, 1).value == student_name:
                        self.sheet.cell(excel_row, 2, value=current_time)
                        self.table_widget.setItem(row, 1, QTableWidgetItem(current_time))
                        break

                # 엑셀 파일 저장
                self.wb.save(self.excel_path)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("맑은 고딕", 10))  # 애플리케이션 전체 기본 폰트 설정
    window = ExcelApp()
    window.show()
    sys.exit(app.exec_())
