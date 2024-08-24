"""
sinsang_file.py 신상명세서 파일 만들기   Ver 1.0_240516
엑셀 신상데이터(신상_자료.xlsx)를 신상명세서.xlsx
형태로 out 폴더에 여러 엑셀 파일로 저장
"""
from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook(filename = "./school/신상_자료.xlsx", data_only = True)
# print(wb.sheetnames)

ws = wb[wb.sheetnames[0]]


tmp = []
for i in range(4, ws.max_row+1):
    tmp.append([])
    for cell in ws[i]:
        tmp[-1].append(cell.value)

#  print(tmp)

for person in tmp:
    wb2 = load_workbook(filename = "./school/invest/양식_신상명세서.xlsx")
    ws2 = wb2[wb2.sheetnames[0]] 
    
    for i in range(1,len(person)+1):
        # ws2.cell(row=5, column = 5).value = person[1]
        ws2['C5'] = person[0]
        ws2['E5'] = person[1]
        ws2['G5'] = person[6]
        ws2['I5'] = person[7]
        ws2['K5'] = person[4]
        ws2['C7'] = person[5]
        ws2['C6'] = person[8]
        ws2['K6'] = person[9]
        ws2['F7'] = person[10]
        ws2['C10'] = person[11]
        ws2['K8'] = person[12]
        ws2['C11'] = person[13]
        ws2['K7'] = person[14]
        ws2['C8'] = person[15]
        ws2['F8'] = person[16]
        ws2['K9'] = person[17]
        ws2['F9'] = person[18]
        ws2['F10'] = person[19]
        ws2['F11'] = person[20]
        ws2['I8'] = person[21]
        ws2['I9'] = person[22]
        ws2['F12'] = person[23]
        ws2['I10'] = person[24]
        ws2['I11'] = person[25]
        ws2['I12'] = person[26]
        ws2['K10'] = person[27]
        ws2['K11'] = person[28]
        ws2['K12'] = person[29]
        ws2['C13'] = person[30]
        ws2['D14'] = person[31]
        ws2['J14'] = person[32]
        ws2['D15'] = person[33]
        ws2['J15'] = person[34]
        ws2['D16'] = person[35]
        ws2['J16'] = person[36]
        ws2['D17'] = person[37]
        ws2['J17'] = person[38]
        ws2['D18'] = person[39]
        ws2['J18'] = person[40]

    wb2.save("./school/invest/"+str(person[0])+"_신상_"+str(person[1])+".xlsx")
    print(str(person[1]))
print("Saving....")
    