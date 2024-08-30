"""
schoolclub_collect.py   동아리 반별 취합   Ver 1.0_240607
"""
import os
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QLabel, QMessageBox, QLineEdit
from openpyxl import Workbook, load_workbook

class SchoolClubCollector(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("학교 클럽 데이터 취합")
        self.setGeometry(200, 200, 400, 250)

        self.select_folder_btn = QPushButton("폴더 선택", self)
        self.select_folder_btn.setGeometry(20, 50, 350, 40)
        self.select_folder_btn.clicked.connect(self.select_folder)

        self.output_filename_label = QLabel("출력 파일명:", self)
        self.output_filename_label.setGeometry(20, 120, 120, 30)

        self.output_filename_edit = QLineEdit("취합_동아리", self) # 기본값 설정
        self.output_filename_edit.setGeometry(100, 120, 270, 30)

        self.result_label = QLabel(self)
        self.result_label.setGeometry(20, 160, 340, 100)

        self.show()

    def select_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "폴더 선택")
        if folder_path:
            self.collect_data(folder_path)

    def collect_data(self, folder_path):
        files = os.listdir(folder_path)
        # 파일명 앞 글자가 자료_ 인 파일 수집
        excel_files = [file for file in files if file.startswith("자료_") and file.lower().endswith('.xlsx')]
        if not excel_files:
            QMessageBox.warning(self, "경고", "선택한 폴더에 적절한 엑셀 파일이 없습니다.")
            return

        output_filename = self.output_filename_edit.text().strip()
        if not output_filename:
            QMessageBox.warning(self, "경고", "출력 파일명을 입력하세요.")
            return

        wb2 = Workbook()
        ws2 = wb2.active
        cnt = 0
        student = 0

        for filename in excel_files:
            wb = load_workbook(os.path.join(folder_path, filename))
            ws = wb.active
            cnt += 1
            if cnt == 1:
                num = 1
            else:
                num = 2
            for row in ws.iter_rows(min_row=num):
                data = []
                student += 1
                for cell in row:
                    data.append(cell.value)
                ws2.append(data)

        ws2.title = '전체'
        # 출력 저장 폳더
        fold_path = "./school/club"
        output_file_path = os.path.join(fold_path, f"{output_filename}.xlsx")
        wb2.save(output_file_path)
        self.result_label.setText(f"총 파일 수: {cnt}, 총 학생 수: {student-1}\n취합된 데이터를 저장했습니다:\n {output_file_path}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SchoolClubCollector()
    sys.exit(app.exec_())
    end_time = time.time()
