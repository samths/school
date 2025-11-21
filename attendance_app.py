"""
attendance_app.py 출석 리포트 프로그램 Ver 1.1_251122
"""
import sqlite3
import datetime
import os
import shutil
from tkinter import *
from tkinter import ttk, messagebox, simpledialog, filedialog
from reportlab.lib.pagesizes import A4, landscape  # landscape 임포트 유지
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl

# --- 설정 상수 ---
DB_FILE = "attendancek.db"
# NanumGothic.ttf 파일은 프로그램 실행 경로에 있어야 합니다.
KOREAN_FONT_NAME = "NanumGothic"
KOREAN_FONT_FILE = "NanumGothic.ttf"

# 출결 상태 코드 및 설명
STATUS_CODES = {
    'P': '출석',
    'A': '결석',
    'L': '병가/지각',
    'T': '훈련/외출',
    'S': '특이사항',
    'W': '주말',  # 보고서용
    'H': '공휴일',  # 보고서용
    '': '미체크',
}


# --- DB 클래스 ---
class DB:
    def __init__(self, dbfile=DB_FILE):
        if hasattr(self, 'conn') and self.conn:
            self.conn.close()
        self.conn = sqlite3.connect(dbfile)
        self.cur = self.conn.cursor()
        self._create_tables()
        self._check_and_add_bigo_column()

    def _create_tables(self):
        self.cur.execute('''
            CREATE TABLE IF NOT EXISTS students(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                roll_no TEXT UNIQUE NOT NULL,
                name TEXT
            )
        ''')
        self.cur.execute('''
            CREATE TABLE IF NOT EXISTS attendance(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                roll_no TEXT,
                date TEXT,
                status TEXT,
                bigo TEXT,  -- bigo 컬럼 추가
                UNIQUE(roll_no,date)
            )
        ''')
        self.cur.execute('''
            CREATE TABLE IF NOT EXISTS holidays(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT UNIQUE NOT NULL,
                name TEXT
            )
        ''')
        # 일일 비고 저장을 위한 별도 테이블 추가
        self.cur.execute('''
            CREATE TABLE IF NOT EXISTS daily_notes(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT UNIQUE NOT NULL,
                note TEXT
            )
        ''')
        self.conn.commit()

    def _check_and_add_bigo_column(self):
        """attendance 테이블에 bigo 컬럼이 없으면 추가"""
        try:
            self.cur.execute("SELECT bigo FROM attendance LIMIT 1")
        except sqlite3.OperationalError:
            self.cur.execute("ALTER TABLE attendance ADD COLUMN bigo TEXT")
            self.conn.commit()
            # print("DB: 'attendance' 테이블에 'bigo' 컬럼을 추가했습니다.")

    def add_student(self, roll_no, name):
        try:
            self.cur.execute("INSERT INTO students(roll_no,name) VALUES(?,?)", (roll_no, name))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def update_student(self, roll_no, name):
        self.cur.execute("UPDATE students SET name=? WHERE roll_no=?", (name, roll_no))
        self.conn.commit()

    def delete_student(self, roll_no):
        self.cur.execute("DELETE FROM students WHERE roll_no=?", (roll_no,))
        self.cur.execute("DELETE FROM attendance WHERE roll_no=?", (roll_no,))
        self.conn.commit()

    def get_all_students(self, order_by="roll_no"):
        self.cur.execute(f"SELECT roll_no,name FROM students ORDER BY {order_by}")
        return self.cur.fetchall()

    def get_student(self, roll_no):
        self.cur.execute("SELECT roll_no,name FROM students WHERE roll_no=?", (roll_no,))
        return self.cur.fetchone()

    def set_attendance(self, roll_no, date_str, status, bigo=''):
        """출석 상태 및 비고 저장/업데이트"""
        # 기존에 bigo가 있는 경우 유지하도록 수정
        # 이 함수는 set_attendance_from_tree 또는 toggle_selected_status에서 호출되므로,
        # bigo는 이미 UI/로직에서 전달받았거나 기존 값을 가져와야 합니다.

        # 1. 기존 bigo를 가져옴 (bigo 파라미터가 비어있을 경우)
        if not bigo:
            self.cur.execute("SELECT bigo FROM attendance WHERE roll_no=? AND date=?", (roll_no, date_str))
            existing_bigo = self.cur.fetchone()
            final_bigo = existing_bigo[0] if existing_bigo and existing_bigo[0] else ''
        else:
            final_bigo = bigo

        # 2. 상태와 최종 bigo를 함께 저장
        self.cur.execute('''
            INSERT INTO attendance(roll_no,date,status,bigo) VALUES(?,?,?,?)
            ON CONFLICT(roll_no,date) DO UPDATE SET status=excluded.status, bigo=excluded.bigo
        ''', (roll_no, date_str, status, final_bigo))
        self.conn.commit()

    def set_daily_note(self, date_str, note):
        """특정 날짜의 '일일 비고'를 daily_notes 테이블에 저장/업데이트"""
        self.cur.execute('''
            INSERT INTO daily_notes(date, note) VALUES(?,?)
            ON CONFLICT(date) DO UPDATE SET note=excluded.note
        ''', (date_str, note))
        self.conn.commit()

    def get_daily_note(self, date_str):
        """특정 날짜의 '일일 비고'를 불러옵니다."""
        self.cur.execute("SELECT note FROM daily_notes WHERE date=?", (date_str,))
        result = self.cur.fetchone()
        return result[0] if result else ""

    def get_student_note(self, roll_no, date_str):
        """특정 학생의 특정 날짜 비고를 불러옵니다."""
        self.cur.execute("SELECT bigo FROM attendance WHERE roll_no=? AND date=?", (roll_no, date_str))
        result = self.cur.fetchone()
        return result[0] if result and result[0] else ""

    def bulk_set_attendance(self, date_str, status):
        students = self.get_all_students()
        for r, _ in students:
            # 일괄 처리 시 기존 비고 유지 (비고 파라미터를 빈 문자열로 전달)
            self.set_attendance(r, date_str, status, bigo='')

    def get_attendance_for_date(self, date_str):
        """roll_no: (status, bigo) 딕셔너리 반환"""
        self.cur.execute("SELECT roll_no,status,bigo FROM attendance WHERE date=?", (date_str,))
        return {r: (s, b if b else '') for r, s, b in self.cur.fetchall()}

    def get_date_range_attendance(self, start_date, end_date):
        """특정 기간의 출석 데이터 조회 (월간 보고서용)"""
        self.cur.execute(
            "SELECT roll_no,date,status FROM attendance WHERE date >= ? AND date <= ?",
            (start_date, end_date)
        )
        rows = self.cur.fetchall()
        data = {}
        for r, d, s in rows:
            data.setdefault(r, {})[d] = s
        return data

    def get_holidays_in_range(self, start_date, end_date):
        """특정 기간의 공휴일 조회"""
        self.cur.execute(
            "SELECT date, name FROM holidays WHERE date >= ? AND date <= ? ORDER BY date",
            (start_date, end_date)
        )
        return self.cur.fetchall()

    def is_holiday(self, date_str):
        """해당 날짜가 공휴일인지 확인"""
        self.cur.execute("SELECT name FROM holidays WHERE date=?", (date_str,))
        result = self.cur.fetchone()
        return result[0] if result else None

    def add_holiday(self, date_str, name):
        """공휴일 추가"""
        try:
            self.cur.execute("INSERT INTO holidays(date, name) VALUES(?, ?)", (date_str, name))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def delete_holiday(self, date_str):
        """공휴일 삭제"""
        self.cur.execute("DELETE FROM holidays WHERE date=?", (date_str,))
        self.conn.commit()

    def get_all_holidays(self):
        """모든 공휴일 조회"""
        self.cur.execute("SELECT date, name FROM holidays ORDER BY date")
        return self.cur.fetchall()

    def close(self):
        if self.conn:
            self.conn.close()


# --- 헬퍼 함수 ---
def is_weekend(date_obj):
    return date_obj.weekday() in (5, 6)  # 토(5), 일(6)


def is_non_working_day(date_obj, db):
    """주말 또는 공휴일인지 확인"""
    if is_weekend(date_obj):
        return True, "주말"
    holiday_name = db.is_holiday(date_obj.isoformat())
    if holiday_name:
        return True, f"공휴일({holiday_name})"
    return False, None


def ensure_roll_format(roll):
    return roll.isdigit() and len(roll) == 5


# --- 메인 클래스 ---
class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("출석 관리 시스템 v1.1")
        self.root.geometry("1100x800")
        self.root.minsize(1000, 650)
        self.db = DB()
        self.style = ttk.Style()
        self._register_korean_font()
        self._setup_style()
        self._build_ui()
        self.refresh_student_list()
        self.update_statistics()  # 선택된 날짜 통계로 로드
        self.load_attendance_for_date()  # 시작 시 오늘 날짜 출석 로드

    def _register_korean_font(self):
        try:
            if os.path.exists(KOREAN_FONT_FILE):
                pdfmetrics.registerFont(TTFont(KOREAN_FONT_NAME, KOREAN_FONT_FILE))
                pdfmetrics.registerFontFamily(KOREAN_FONT_NAME, normal=KOREAN_FONT_NAME, bold=KOREAN_FONT_NAME)
                self.korean_styles = getSampleStyleSheet()
                self.korean_styles['Normal'].fontName = KOREAN_FONT_NAME
                self.korean_styles['Normal'].fontSize = 10
                self.korean_styles['Heading1'].fontName = KOREAN_FONT_NAME
                self.korean_styles['Heading2'].fontName = KOREAN_FONT_NAME
                self.korean_styles['Heading3'].fontName = KOREAN_FONT_NAME
            else:
                messagebox.showwarning("폰트 경고",
                                       f"{KOREAN_FONT_FILE} 파일이 없습니다.\nPDF 한글 출력이 제한될 수 있습니다.")
                self.korean_styles = getSampleStyleSheet()  # 기본 폰트 사용
        except Exception as e:
            messagebox.showwarning("폰트 오류", f"폰트 로드 중 오류: {e}")
            self.korean_styles = getSampleStyleSheet()

    def _setup_style(self):
        self.style.theme_use('clam')
        self.style.configure("TFrame", background="#f6f8ff")
        self.style.configure("Header.TLabel", font=("맑은 고딕", 18, "bold"),
                             background="#3a6df0", foreground="white", padding=10)
        self.style.configure("TButton", font=("맑은 고딕", 10, "bold"), padding=6)
        self.style.map("TButton", foreground=[('!disabled', 'white')],
                       background=[('!disabled', '#5b8cff')])
        self.style.configure("Accent.TButton", background="#ff7ab6", foreground="white")
        self.style.configure("Success.TButton", background="#28a745", foreground="white")
        self.style.configure("Danger.TButton", background="#dc3545", foreground="white")
        self.style.configure("Treeview", font=("맑은 고딕", 11), rowheight=28)
        self.style.configure("Stats.TLabel", font=("맑은 고딕", 12, "bold"),
                             background="#ffffff", padding=10)

    def _build_ui(self):
        # Header
        header = ttk.Frame(self.root)
        header.pack(fill=X)
        ttk.Label(header, text="출결 관리 시스템 v1.1", style="Header.TLabel").pack(fill=X)

        # 메뉴바 추가
        menubar = Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="파일", menu=file_menu)
        file_menu.add_command(label="백업", command=self.backup_database)
        file_menu.add_command(label="복원", command=self.restore_database)
        file_menu.add_separator()
        file_menu.add_command(label="엑셀 내보내기", command=self.export_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="종료", command=self.on_close)

        # 공휴일 메뉴 추가
        holiday_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="공휴일 관리", menu=holiday_menu)
        holiday_menu.add_command(label="공휴일 추가", command=self.add_holiday_dialog)
        holiday_menu.add_command(label="공휴일 목록", command=self.view_holidays_dialog)
        holiday_menu.add_command(label="📥 엑셀로 공휴일 불러오기", command=self.import_holidays_from_excel)

        # 통계 프레임 (상단)
        stats_frame = ttk.LabelFrame(self.root, text="통계 대시보드", padding=10)
        stats_frame.pack(fill=X, padx=10, pady=5)

        self.stats_labels = {}
        stats_inner = ttk.Frame(stats_frame)
        stats_inner.pack(fill=X)

        # 통계 표시 라벨들
        self.stats_labels['total'] = ttk.Label(stats_inner, text="총 학생: 0",
                                               style="Stats.TLabel")
        self.stats_labels['total'].pack(side=LEFT, padx=20)

        self.stats_labels['today_present'] = ttk.Label(stats_inner, text="출석: 0",
                                                       style="Stats.TLabel", foreground="blue")
        self.stats_labels['today_present'].pack(side=LEFT, padx=20)

        self.stats_labels['today_absent'] = ttk.Label(stats_inner, text="결석: 0",
                                                      style="Stats.TLabel", foreground="red")
        self.stats_labels['today_absent'].pack(side=LEFT, padx=20)

        self.stats_labels['note'] = ttk.Label(stats_inner, text="일일 비고: ",
                                              style="Stats.TLabel", foreground="green")
        self.stats_labels['note'].pack(side=LEFT, padx=20, fill=X, expand=True)

        # 비고 입력 버튼 추가
        ttk.Button(stats_inner, text="📝 비고 입력",
                   command=self.input_daily_note).pack(side=RIGHT, padx=5)

        ttk.Button(stats_inner, text="통계 새로고침",
                   command=self.update_statistics).pack(side=RIGHT, padx=5)

        # Main frames
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill=BOTH, expand=True)

        left = ttk.Frame(main)
        left.pack(side=LEFT, fill=Y, padx=(0, 8))

        right = ttk.Frame(main)
        right.pack(side=RIGHT, fill=BOTH, expand=True)

        # --- Left: Student management ---
        sframe = ttk.LabelFrame(left, text="학생 관리", padding=8)
        sframe.pack(fill=Y, expand=True)

        # 검색
        search_frame = ttk.Frame(sframe)
        search_frame.pack(fill=X, padx=6, pady=6)
        ttk.Label(search_frame, text="검색:").pack(side=LEFT)
        self.search_var = StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side=LEFT, fill=X, expand=True, padx=(5, 0))
        search_entry.bind("<KeyRelease>", lambda e: self.refresh_student_list())

        # 버튼들
        btns = ttk.Frame(sframe)
        btns.pack(fill=X, padx=6, pady=4)
        ttk.Button(btns, text="추가", command=self.add_student_dialog,
                   style="Success.TButton").pack(side=LEFT, fill=X, expand=True, padx=2)
        ttk.Button(btns, text="수정", command=self.edit_selected).pack(
            side=LEFT, fill=X, expand=True, padx=2)
        ttk.Button(btns, text="삭제", command=self.delete_selected,
                   style="Danger.TButton").pack(side=LEFT, fill=X, expand=True, padx=2)

        bulk = ttk.Frame(sframe)
        bulk.pack(fill=X, padx=6, pady=4)
        ttk.Button(bulk, text="학번 일괄 추가 (10101-10136)",
                   command=self.bulk_generate_rolls).pack(fill=X)

        # 공휴일 관리 버튼 추가 (메인 화면에서 바로 접근)
        holiday_btn_frame = ttk.Frame(sframe)
        holiday_btn_frame.pack(fill=X, padx=6, pady=6)
        ttk.Button(holiday_btn_frame, text="🏖️ 공휴일 목록 관리",
                   command=self.view_holidays_dialog,
                   style="Accent.TButton").pack(fill=X)

        # Student tree with scrollbar
        tree_frame = ttk.Frame(sframe)
        tree_frame.pack(fill=BOTH, expand=True, padx=6, pady=(4, 0))

        self.student_tree = ttk.Treeview(tree_frame, columns=("roll", "name"),
                                         show="headings", height=18)
        self.student_tree.heading("roll", text="학번")
        self.student_tree.column("roll", width=80, anchor=CENTER)
        self.student_tree.heading("name", text="이름")
        self.student_tree.column("name", width=150, anchor=W)

        scrollbar_s = ttk.Scrollbar(tree_frame, orient=VERTICAL,
                                    command=self.student_tree.yview)
        self.student_tree.configure(yscroll=scrollbar_s.set)

        self.student_tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar_s.pack(side=RIGHT, fill=Y)
        self.student_tree.bind("<ButtonRelease-1>", self.load_attendance_for_date)

        # --- Right: Attendance & Reports ---
        top_right = ttk.Frame(right)
        top_right.pack(fill=X)

        datefrm = ttk.Frame(top_right)
        datefrm.pack(side=LEFT, padx=6, pady=6)
        ttk.Label(datefrm, text="날짜:").pack(side=LEFT)
        self.date_var = StringVar()
        self.date_var.set(datetime.date.today().isoformat())
        self.date_entry = ttk.Entry(datefrm, textvariable=self.date_var, width=12)
        self.date_entry.pack(side=LEFT, padx=(6, 0))
        ttk.Button(datefrm, text="오늘", command=self.set_today).pack(side=LEFT, padx=6)
        # 날짜 변경 시 통계 및 출석 로드
        self.date_entry.bind("<Return>", lambda e: (self.load_attendance_for_date(), self.update_statistics()))
        self.date_var.trace_add("write", lambda name, index, mode: self.update_statistics())

        # 일괄 처리 버튼
        quick = ttk.Frame(top_right)
        quick.pack(side=RIGHT, padx=6)
        ttk.Button(quick, text="전체 출석(P)", style="Success.TButton",
                   command=lambda: self.bulk_mark("P")).pack(side=LEFT, padx=2)
        ttk.Button(quick, text="전체 결석(A)", style="Danger.TButton",
                   command=lambda: self.bulk_mark("A")).pack(side=LEFT, padx=2)
        ttk.Button(quick, text="전체 지각/병가(L)", style="Accent.TButton",
                   command=lambda: self.bulk_mark("L")).pack(side=LEFT, padx=2)

        # Attendance list
        att_frame = ttk.LabelFrame(right, text="출석 체크 (더블클릭: P→A→L→T→S→빈칸)", padding=8)
        att_frame.pack(fill=BOTH, expand=True, padx=6, pady=6)

        # bigo 컬럼 추가 (화면에는 미표시, 로직에서 사용)
        cols = ("roll", "name", "status", "bigo")
        self.att_tree = ttk.Treeview(att_frame, columns=cols, show="headings")
        self.att_tree.heading("roll", text="학번")
        self.att_tree.column("roll", width=80, anchor=CENTER)
        self.att_tree.heading("name", text="이름")
        self.att_tree.column("name", width=120, anchor=W)
        self.att_tree.heading("status", text="출결 (P/A/L/T/S)")
        self.att_tree.column("status", width=120, anchor=CENTER)
        # bigo 컬럼 숨김
        self.att_tree.column("bigo", width=0, stretch=NO)

        scrollbar = ttk.Scrollbar(att_frame, orient=VERTICAL, command=self.att_tree.yview)
        self.att_tree.configure(yscroll=scrollbar.set)

        self.att_tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.att_tree.bind("<Double-1>", self.toggle_selected_status)  # 상태 토글
        self.att_tree.bind("<Button-3>", self.show_context_menu)  # 우클릭 메뉴

        # 하단 버튼들
        rside = ttk.Frame(right)
        rside.pack(fill=X, padx=6, pady=5)
        ttk.Button(rside, text="출석 로드 (날짜/학생)", command=self.load_attendance_for_date).pack(
            side=LEFT, padx=2)
        ttk.Button(rside, text="출석 저장", style="Success.TButton",
                   command=self.save_attendance_from_tree).pack(side=LEFT, padx=2)

        ttk.Separator(rside, orient=VERTICAL).pack(side=LEFT, padx=10, fill=Y)

        ttk.Button(rside, text="일일 PDF 보고서", command=self.export_daily_pdf, style="Accent.TButton").pack(
            side=LEFT, padx=2)
        ttk.Button(rside, text="월간 PDF 보고서", command=self.export_monthly_pdf, style="Accent.TButton").pack(
            side=LEFT, padx=2)

        # Status bar
        self.status_var = StringVar()
        statusbar = ttk.Label(self.root, textvariable=self.status_var,
                              relief=SUNKEN, anchor=W)
        statusbar.pack(side=BOTTOM, fill=X)
        self.set_status("준비 완료 - 출석 관리 시스템 v1.4 (완성)")

    def set_status(self, msg):
        self.status_var.set(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")
        self.root.update_idletasks()

    def on_close(self):
        self.db.close()
        self.root.destroy()

    def show_context_menu(self, event):
        """출석 목록에서 우클릭 시 비고 입력 메뉴 표시"""
        selected_item = self.att_tree.identify_row(event.y)
        if not selected_item:
            return

        self.att_tree.selection_set(selected_item)  # 우클릭 시 선택되도록 함

        try:
            # 롤 번호를 가져옴
            roll_no = self.att_tree.item(selected_item)['values'][0]
            name = self.att_tree.item(selected_item)['values'][1]

            menu = Menu(self.root, tearoff=0)
            menu.add_command(label=f"{name} ({roll_no}) 비고 입력",
                             command=lambda: self.input_student_note(roll_no))

            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def input_student_note(self, roll_no):
        """특정 학생의 비고를 입력받아 저장"""
        date_str = self.date_var.get()
        student = self.db.get_student(roll_no)
        name = student[1] if student else roll_no
        current_note = self.db.get_student_note(roll_no, date_str)

        new_note = simpledialog.askstring("학생 비고 입력",
                                          f"{date_str}의 {name} ({roll_no}) 비고를 입력하세요:",
                                          initialvalue=current_note,
                                          parent=self.root)

        if new_note is not None:
            # 상태는 그대로 두고 bigo만 업데이트합니다.
            att_data = self.db.get_attendance_for_date(date_str)
            current_status = att_data.get(roll_no, ('', ''))[0]

            self.db.cur.execute('''
                INSERT INTO attendance(roll_no,date,status,bigo) VALUES(?,?,?,?)
                ON CONFLICT(roll_no,date) DO UPDATE SET bigo=excluded.bigo, status=excluded.status
            ''', (roll_no, date_str, current_status, new_note.strip()))
            self.db.conn.commit()

            self.load_attendance_for_date()
            self.set_status(f"{date_str} {name} 비고 저장 완료: {new_note.strip()}")

    # ==========================================================
    # 학생/출석 관리 로직
    # ==========================================================

    def update_statistics(self):
        """선택된 날짜의 통계 대시보드 새로고침"""
        date_str = self.date_var.get()
        try:
            date_obj = datetime.date.fromisoformat(date_str)
        except ValueError:
            self.set_status("오류: 날짜 형식이 올바르지 않습니다.")
            return

        total_students = len(self.db.get_all_students())
        # att_data는 (status, bigo)를 반환
        att_data = self.db.get_attendance_for_date(date_str)

        is_non_work, reason = is_non_working_day(date_obj, self.db)

        present_codes = ('P', 'T', 'L')
        today_present = sum(1 for status, _ in att_data.values() if status in present_codes)

        self.stats_labels['total'].config(text=f"총 학생: {total_students}")

        if is_non_work:
            note = f"⚠️ {date_str} - 비근무일 ({reason})"
            self.stats_labels['today_present'].config(text="출석: N/A", foreground="gray")
            self.stats_labels['today_absent'].config(text="결석: N/A", foreground="gray")
            self.stats_labels['note'].config(text=note, foreground="green")
        else:
            today_absent = total_students - today_present
            self.stats_labels['today_present'].config(text=f"출석: {today_present}", foreground="blue")
            self.stats_labels['today_absent'].config(text=f"결석: {today_absent}", foreground="red")

            # DB에서 해당 날짜의 일일 비고를 조회
            daily_note = self.db.get_daily_note(date_str)

            display_note = daily_note if len(daily_note) <= 30 else daily_note[:27] + "..."
            self.stats_labels['note'].config(text=f"일일 비고: {display_note}", foreground="green")

        self.set_status(f"{date_str} 통계 새로고침 완료")

    def load_attendance_for_date(self, event=None):
        """날짜에 해당하는 출석 데이터를 로드하여 Treeview에 표시"""
        date_str = self.date_var.get()
        try:
            date_obj = datetime.date.fromisoformat(date_str)
        except ValueError:
            messagebox.showerror("오류", "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD).")
            return

        for item in self.att_tree.get_children():
            self.att_tree.delete(item)

        is_non_work, reason = is_non_working_day(date_obj, self.db)
        students = self.db.get_all_students()
        # att_data는 (status, bigo)를 반환
        att_data = self.db.get_attendance_for_date(date_str)

        for roll, name in students:
            status, bigo = att_data.get(roll, ('', ''))

            if is_non_work:
                status_code = 'H' if date_obj.weekday() < 5 else 'W'
                display_status = STATUS_CODES.get(status_code, status_code)
                # 비고는 저장된 값 또는 비근무일 사유를 사용
                final_bigo = bigo if bigo else reason
            else:
                display_status = STATUS_CODES.get(status, status)
                final_bigo = bigo

            # bigo 컬럼을 추가하여 Treeview에 로드 (숨겨진 컬럼)
            self.att_tree.insert("", "end", values=(roll, name, display_status, final_bigo), tags=(status,))

        if is_non_work:
            self.set_status(f"⚠️ {date_str} - {reason}입니다. 출석 기록은 저장되지 않습니다.")
        else:
            self.set_status(f"{date_str} 출석 목록 로드 완료.")

        # 통계도 함께 새로고침
        self.update_statistics()

    def save_attendance_from_tree(self):
        """Treeview의 출석 데이터를 DB에 저장 (bigo 포함)"""
        date_str = self.date_var.get()
        try:
            date_obj = datetime.date.fromisoformat(date_str)
        except ValueError:
            messagebox.showerror("오류", "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD).")
            return

        is_non_work, _ = is_non_working_day(date_obj, self.db)
        if is_non_work:
            messagebox.showwarning("경고", f"{date_str}은 비근무일입니다. 출석 데이터를 저장할 수 없습니다.")
            return

        updated_count = 0

        for item in self.att_tree.get_children():
            # bigo 컬럼도 함께 가져옴
            roll_no, _, display_status, bigo = self.att_tree.item(item)['values']
            status_code = next((k for k, v in STATUS_CODES.items() if v == display_status), display_status)

            if status_code not in ('W', 'H'):
                # bigo를 포함하여 저장
                self.db.set_attendance(roll_no, date_str, status_code, bigo=bigo)
                updated_count += 1

        self.load_attendance_for_date()
        self.update_statistics()
        self.set_status(f"{date_str} 출석 기록 {updated_count}건 저장 완료.")

    def toggle_selected_status(self, event):
        """선택된 학생의 출결 상태를 순환 변경"""
        selected_item = self.att_tree.selection()
        if not selected_item:
            return

        date_str = self.date_var.get()
        date_obj = datetime.date.fromisoformat(date_str)
        is_non_work, _ = is_non_working_day(date_obj, self.db)

        if is_non_work:
            messagebox.showwarning("경고", f"{date_str}은 비근무일입니다. 상태를 변경할 수 없습니다.")
            return

        STATUS_CYCLE = ['P', 'A', 'L', 'T', 'S', '']

        item = selected_item[0]
        # bigo도 가져와서 유지
        roll_no, name, current_display_status, bigo = self.att_tree.item(item)['values']
        current_status = next((k for k, v in STATUS_CODES.items() if v == current_display_status), '')

        try:
            current_index = STATUS_CYCLE.index(current_status)
            new_status_code = STATUS_CYCLE[(current_index + 1) % len(STATUS_CYCLE)]
        except ValueError:
            new_status_code = 'P'

        new_display_status = STATUS_CODES.get(new_status_code, new_status_code)

        # bigo는 유지하고 새로운 상태로 Treeview 업데이트
        self.att_tree.item(item, values=(roll_no, name, new_display_status, bigo), tags=(new_status_code,))

        # bigo를 포함하여 DB에 저장
        self.db.set_attendance(roll_no, date_str, new_status_code, bigo=bigo)
        self.update_statistics()
        self.set_status(f"{name} ({roll_no}) 상태 변경: {new_display_status} 저장 완료.")

    def input_daily_note(self):
        """선택된 날짜에 대한 일일 비고를 입력받아 저장"""
        date_str = self.date_var.get()
        current_note = self.db.get_daily_note(date_str)

        note_win = Toplevel(self.root)
        note_win.title(f"{date_str} 일일 비고 입력")
        note_win.geometry("500x300")

        ttk.Label(note_win, text=f"날짜: {date_str}\n일일 비고 사항을 입력하세요:",
                  justify=LEFT).pack(padx=10, pady=10, fill=X)

        note_text = Text(note_win, height=8, width=50)
        note_text.insert(END, current_note)
        note_text.pack(padx=10, pady=5, fill=BOTH, expand=True)

        def save_note():
            note_content = note_text.get("1.0", END).strip()
            self.db.set_daily_note(date_str, note_content)

            messagebox.showinfo("저장 완료", f"{date_str} 일일 비고가 저장되었습니다.")
            self.set_status(f"{date_str} 일일 비고 저장 완료.")
            self.update_statistics()
            note_win.destroy()

        ttk.Button(note_win, text="저장", command=save_note, style="Success.TButton").pack(
            side=RIGHT, padx=10, pady=10)
        ttk.Button(note_win, text="취소", command=note_win.destroy).pack(
            side=RIGHT, padx=10, pady=10)

    def set_today(self):
        self.date_var.set(datetime.date.today().isoformat())
        self.load_attendance_for_date()

    def bulk_mark(self, status):
        """전체 학생에게 일괄적으로 출결 상태를 지정"""
        date_str = self.date_var.get()
        try:
            date_obj = datetime.date.fromisoformat(date_str)
        except ValueError:
            messagebox.showerror("오류", "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD).")
            return

        is_non_work, _ = is_non_working_day(date_obj, self.db)
        if is_non_work:
            messagebox.showwarning("경고", f"{date_str}은 비근무일이므로 일괄 처리를 할 수 없습니다.")
            return

        # bigo는 기존 값 유지하며 일괄 저장
        self.db.bulk_set_attendance(date_str, status)
        self.load_attendance_for_date()
        self.update_statistics()
        self.set_status(f"{date_str} 날짜에 모든 학생에게 상태 '{STATUS_CODES.get(status, status)}' 저장 완료")

    def refresh_student_list(self):
        """학생 목록 트리뷰를 새로고침 (검색 기능 포함)"""
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)

        search_term = self.search_var.get().strip().lower()
        students = self.db.get_all_students()

        filtered_students = []
        for roll, name in students:
            if search_term in roll.lower() or search_term in name.lower():
                filtered_students.append((roll, name))

        for roll, name in filtered_students:
            self.student_tree.insert("", "end", values=(roll, name))

    def add_student_dialog(self):
        """학생 추가 대화상자"""
        roll_no = simpledialog.askstring("학생 추가", "학번(5자리 숫자)을 입력하세요:", parent=self.root)
        if not roll_no:
            return

        if not ensure_roll_format(roll_no):
            messagebox.showerror("오류", "학번은 5자리 숫자여야 합니다.")
            return

        name = simpledialog.askstring("학생 추가", "이름을 입력하세요:", parent=self.root)
        if not name:
            return

        if self.db.add_student(roll_no, name):
            messagebox.showinfo("성공", f"학생 {roll_no} - {name}이(가) 추가되었습니다.")
            self.refresh_student_list()
            self.load_attendance_for_date()
            self.update_statistics()
        else:
            messagebox.showerror("오류", f"학번 {roll_no}은(는) 이미 존재합니다.")

    def edit_selected(self):
        """선택된 학생 정보 수정"""
        selected_item = self.student_tree.selection()
        if not selected_item:
            messagebox.showinfo("선택", "수정할 학생을 선택하세요.")
            return

        roll_no = self.student_tree.item(selected_item[0])['values'][0]
        current_name = self.student_tree.item(selected_item[0])['values'][1]

        new_name = simpledialog.askstring("학생 정보 수정",
                                          f"학번 {roll_no}의 새 이름을 입력하세요:",
                                          initialvalue=current_name,
                                          parent=self.root)
        if new_name and new_name.strip() != current_name:
            self.db.update_student(roll_no, new_name.strip())
            self.refresh_student_list()
            self.load_attendance_for_date()
            self.set_status(f"학생 {roll_no} 이름이 {new_name.strip()}으로 수정되었습니다.")

    def delete_selected(self):
        """선택된 학생 삭제"""
        selected_item = self.student_tree.selection()
        if not selected_item:
            messagebox.showinfo("선택", "삭제할 학생을 선택하세요.")
            return

        roll_no = self.student_tree.item(selected_item[0])['values'][0]
        name = self.student_tree.item(selected_item[0])['values'][1]

        if messagebox.askyesno("삭제 확인", f"학생 {roll_no} - {name}을(를) 정말 삭제하시겠습니까?\n(출석 기록도 함께 삭제됩니다)"):
            self.db.delete_student(roll_no)
            self.refresh_student_list()
            self.load_attendance_for_date()
            self.update_statistics()
            self.set_status(f"학생 {roll_no} 삭제 완료.")

    def bulk_generate_rolls(self):
        """10101부터 10136까지 학번을 일괄 추가"""
        if not messagebox.askyesno("확인", "학번 10101부터 10136까지 일괄 추가하시겠습니까?\n(이미 존재하는 학번은 건너뜁니다)"):
            return

        added_count = 0
        skipped_count = 0

        for roll_num in range(10101, 10137):
            roll_no = str(roll_num)
            name = f"학생 {roll_num}"

            if self.db.add_student(roll_no, name):
                added_count += 1
            else:
                skipped_count += 1

        messagebox.showinfo("일괄 추가 완료",
                            f"총 {added_count}명의 학생이 추가되었습니다.\n"
                            f"(중복되어 건너뛴 학생: {skipped_count}명)")
        self.refresh_student_list()
        self.load_attendance_for_date()
        self.update_statistics()
        self.set_status(f"학번 10101-10136 일괄 추가 완료 (추가: {added_count}, 건너뜀: {skipped_count})")

    # ==========================================================
    # 파일 관리 및 보고서 출력
    # ==========================================================

    def backup_database(self):
        """데이터베이스 파일 백업"""
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = f"backup_{timestamp}_{DB_FILE}"

            self.db.conn.close()

            shutil.copyfile(DB_FILE, backup_file)

            self.db = DB()

            messagebox.showinfo("백업 성공", f"데이터베이스가 '{backup_file}'로 백업되었습니다.")
            self.set_status(f"DB 백업 완료: {backup_file}")
        except FileNotFoundError:
            messagebox.showerror("오류", f"원본 파일 '{DB_FILE}'을 찾을 수 없습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"백업 중 오류 발생: {e}")
            self.db = DB()

    def restore_database(self):
        """데이터베이스 파일 복원"""
        file_path = filedialog.askopenfilename(
            defaultextension=".db",
            filetypes=[("Database files", "*.db"), ("All files", "*.*")],
            title="복원할 백업 파일을 선택하세요"
        )
        if not file_path:
            return

        if messagebox.askyesno("복원 확인",
                               f"현재 데이터베이스를 '{file_path}' 파일로 덮어씁니다. 진행하시겠습니까?"):
            try:
                self.db.conn.close()
                shutil.copyfile(file_path, DB_FILE)
                self.db = DB()
                messagebox.showinfo("복원 성공", "데이터베이스가 성공적으로 복원되었습니다. 프로그램을 재시작하는 것이 좋습니다.")
                self.set_status(f"DB 복원 완료: {file_path}")
                self.refresh_student_list()
                self.update_statistics()
            except Exception as e:
                messagebox.showerror("복원 오류", f"데이터베이스 복원 중 오류 발생: {e}")
                self.db = DB()

    def export_to_excel(self):
        """전체 데이터를 Excel로 내보내기 (학생 목록 및 기간별 출석)"""
        if 'openpyxl' not in globals():
            messagebox.showerror("오류", "'openpyxl' 라이브러리가 필요합니다.")
            return

        today = datetime.date.today()
        # 기본 시작 날짜를 지난 달 1일로 설정
        initial_start = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1).isoformat()

        start_date = simpledialog.askstring("Excel 내보내기",
                                            "출석 기록 시작 날짜 (YYYY-MM-DD):",
                                            initialvalue=initial_start,
                                            parent=self.root)
        if not start_date: return
        end_date = simpledialog.askstring("Excel 내보내기",
                                          "출석 기록 종료 날짜 (YYYY-MM-DD):",
                                          initialvalue=today.isoformat(),
                                          parent=self.root)
        if not end_date: return

        try:
            start_date_obj = datetime.date.fromisoformat(start_date)
            end_date_obj = datetime.date.fromisoformat(end_date)
        except ValueError:
            messagebox.showerror("오류", "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD).")
            return

        if start_date_obj > end_date_obj:
            messagebox.showerror("오류", "시작 날짜가 종료 날짜보다 늦을 수 없습니다.")
            return

        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="데이터를 저장할 Excel 파일 경로를 지정하세요"
            )
            if not file_path:
                return

            self.set_status("Excel 보고서 생성 중...")

            wb = openpyxl.Workbook()
            ws_summary = wb.active
            ws_summary.title = "학생_출석요약"

            # 1. 출석 요약 시트 준비
            students = self.db.get_all_students()
            # 월간 보고서와 동일하게 status만 가져옴
            att_data = self.db.get_date_range_attendance(start_date, end_date)

            # 날짜 리스트 생성
            date_list = []
            current_date = start_date_obj
            while current_date <= end_date_obj:
                date_list.append(current_date.isoformat())
                current_date += datetime.timedelta(days=1)

            # 헤더
            summary_header = ["학번", "이름", "총 일수", "출석(P)", "결석(A)", "지각/병가(L)", "훈련/외출(T)", "특이사항(S)", "미체크()", "주말(W)", "공휴일(H)"]
            ws_summary.append(summary_header)

            # 데이터 채우기
            for roll, name in students:
                stats = {'P': 0, 'A': 0, 'L': 0, 'T': 0, 'S': 0, '': 0, 'W': 0, 'H': 0}
                total_days = len(date_list)

                for date_str in date_list:
                    date_obj = datetime.date.fromisoformat(date_str)
                    is_non_work, reason = is_non_working_day(date_obj, self.db)

                    status = att_data.get(roll, {}).get(date_str, '')

                    if is_non_work:
                        stats['W' if reason == "주말" else 'H'] += 1
                    else:
                        stats[status] += 1

                row_data = [roll, name, total_days] + [stats[k] for k in ['P', 'A', 'L', 'T', 'S', '', 'W', 'H']]
                ws_summary.append(row_data)

            # 2. 상세 출석 시트 준비
            ws_detail = wb.create_sheet("상세_출석기록")

            # 상세 시트에서는 bigo도 포함하도록 att_data를 재조회 (비고 컬럼 추가)
            self.db.cur.execute(
                "SELECT roll_no,date,status,bigo FROM attendance WHERE date >= ? AND date <= ?",
                (start_date, end_date)
            )
            rows = self.db.cur.fetchall()
            detail_att_data = {}
            for r, d, s, b in rows:
                detail_att_data.setdefault(r, {})[d] = (s, b)

            detail_header = ["학번", "이름"]
            for date_str in date_list:
                detail_header.append(date_str)
                detail_header.append(f"{date_str}_비고")

            ws_detail.append(detail_header)

            for roll, name in students:
                row_data = [roll, name]
                for date_str in date_list:
                    date_obj = datetime.date.fromisoformat(date_str)
                    is_non_work, reason = is_non_working_day(date_obj, self.db)

                    status, bigo = detail_att_data.get(roll, {}).get(date_str, ('', ''))

                    if is_non_work:
                        status = 'H' if date_obj.weekday() < 5 else 'W'
                        bigo = bigo if bigo else reason  # DB에 저장된 비고가 없으면 사유 표시

                    row_data.append(status)
                    row_data.append(bigo)
                ws_detail.append(row_data)

            wb.save(file_path)

            messagebox.showinfo("Excel 내보내기 성공",
                                f"데이터가 '{file_path}'에 성공적으로 저장되었습니다.\n기간: {start_date} ~ {end_date}")
            self.set_status(f"Excel 내보내기 완료: {file_path}")

        except Exception as e:
            messagebox.showerror("오류", f"Excel 내보내기 중 오류 발생: {e}")
            self.set_status(f"Excel 오류: {e}")

    def export_daily_pdf(self):
        """일일 출석 현황 PDF로 내보내기 (날짜를 사용자에게 입력 받음)"""
        date_to_report = simpledialog.askstring("일일 PDF 보고서",
                                                "보고서를 출력할 날짜를 입력하세요 (YYYY-MM-DD):",
                                                initialvalue=self.date_var.get(),
                                                parent=self.root)

        if not date_to_report:
            return

        date_str = date_to_report

        try:
            date_obj = datetime.date.fromisoformat(date_str)
            date_display = date_obj.strftime("%Y년 %m월 %d일")
        except ValueError:
            messagebox.showerror("오류", "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD).")
            return

        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="일일 출석 보고서를 저장할 경로를 지정하세요",
                initialfile=f"일일출석보고서_{date_str}.pdf"
            )
            if not file_path:
                return

            self.set_status("일일 PDF 보고서 생성 중...")

            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = self.korean_styles
            Story = []

            # 제목
            Story.append(Paragraph(f"{date_display} 일일 출석 현황 보고서", styles['Heading1']))
            Story.append(Spacer(1, 6))

            # 일일 비고 표시
            daily_note = self.db.get_daily_note(date_str)
            if daily_note:
                Story.append(Paragraph(f"**[일일 비고]** {daily_note}", styles['Normal']))
                Story.append(Spacer(1, 6))

            # 공휴일/주말 체크
            is_non_work, reason = is_non_working_day(date_obj, self.db)
            if is_non_work:
                Story.append(Paragraph(f"⚠️ **비고: {date_display}은(는) {reason}입니다.**", styles['Normal']))
                Story.append(Spacer(1, 12))

            # 데이터 로드 (status, bigo)
            students = self.db.get_all_students()
            att_data = self.db.get_attendance_for_date(date_str)

            data = [["학번", "이름", "출결 상태", "비고"]]

            for roll, name in students:
                status, bigo = att_data.get(roll, ('', ''))

                if is_non_work:
                    status_code = 'H' if date_obj.weekday() < 5 else 'W'
                    display_status = STATUS_CODES.get(status_code, status_code)
                    # 비고는 저장된 값 또는 비근무일 사유를 사용
                    final_bigo = bigo if bigo else reason
                else:
                    display_status = STATUS_CODES.get(status, status)
                    final_bigo = bigo

                row_data = [roll, name, display_status, final_bigo]
                data.append(row_data)

            table_style = TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), KOREAN_FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DDDDDD')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])

            col_widths = [70, 100, 80, 200]
            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)
            Story.append(table)
            Story.append(Spacer(1, 12))

            legend_text = "출결 상태 범례: " + ", ".join([f"{k}: {v}" for k, v in STATUS_CODES.items() if k not in ('W', 'H')]) + ", W: 주말, H: 공휴일"
            Story.append(Paragraph(legend_text, styles['Normal']))

            doc.build(Story)

            messagebox.showinfo("PDF 출력 성공", f"일일 출석 보고서가 '{file_path}'에 성공적으로 저장되었습니다.")
            self.set_status(f"{date_str} 일일 PDF 보고서 출력 완료: {file_path}")

        except Exception as e:
            messagebox.showerror("오류", f"일일 PDF 보고서 생성 중 오류 발생: {e}")
            self.set_status(f"PDF 오류: {e}")

    def export_monthly_pdf(self):
        """월간 출석 통계 PDF로 내보내기 (월간 누계만 출력)"""
        today = datetime.date.today()
        month = simpledialog.askinteger("월간 PDF",
                                        "보고서를 만들 월(Month)을 숫자로 입력하세요:",
                                        initialvalue=today.month,
                                        minvalue=1, maxvalue=12, parent=self.root)
        if month is None: return

        year = simpledialog.askinteger("월간 PDF",
                                       "보고서를 만들 연도(Year)를 숫자로 입력하세요:",
                                       initialvalue=today.year,
                                       minvalue=2000, maxvalue=today.year + 1, parent=self.root)
        if year is None: return

        # 총 일수 입력은 월간 누계를 계산하는 데 필요하므로 유지합니다.
        try:
            if month == 12:
                last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                last_day = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
            default_days = last_day.day
        except ValueError:
            default_days = 30

            # 총 일수(근무일수)를 입력받는 부분은 그대로 유지합니다.
        total_work_days = simpledialog.askinteger("월간 PDF",
                                                  "출석할 **총 일수(근무일수)**를 입력하세요:",
                                                  initialvalue=default_days,
                                                  minvalue=1, maxvalue=31, parent=self.root)
        if total_work_days is None: return

        try:
            start_date = datetime.date(year, month, 1)
            if month == 12:
                end_date = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                end_date = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)

            report_period = f"{year}년 {month}월 01일 ~ {end_date.day}일"
        except ValueError:
            messagebox.showerror("오류", "유효하지 않은 날짜입니다.")
            return

        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="월간 출석 보고서를 저장할 경로를 지정하세요",
                initialfile=f"월간출석누계보고서_{year}년_{month}월.pdf"
            )
            if not file_path:
                return

            self.set_status("월간 PDF 보고서 생성 중...")

            # ----------------------------------------------------
            # PDF 문서 설정 (가로 방향 A4 유지)
            # ----------------------------------------------------
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
            styles = self.korean_styles
            Story = []

            # 제목
            Story.append(Paragraph(f"월간 출석 **누계 통계 보고서** ({report_period})", styles['Heading1']))
            Story.append(Paragraph(f"**기준 출석일:** {total_work_days}일", styles['Normal']))
            Story.append(Spacer(1, 12))

            # ----------------------------------------------------
            # 데이터 집계 로직
            # ----------------------------------------------------
            students = self.db.get_all_students()
            att_data = self.db.get_date_range_attendance(start_date.isoformat(), end_date.isoformat())

            # 날짜 리스트 생성 (기간 내 모든 날짜를 순회하며 통계 계산)
            date_list = []
            current_date = start_date
            while current_date <= end_date:
                date_list.append(current_date)
                current_date += datetime.timedelta(days=1)

            # 테이블 헤더: 학번, 이름 + 통계 항목
            header = ["학번", "이름", "총 일수", "출석(P)", "결석(A)", "지각/병가(L)", "훈련/외출(T)", "특이사항(S)", "미체크()", "주말(W)", "공휴일(H)"]
            data = [header]

            # 데이터 채우기 (월간 누계만 계산)
            for roll, name in students:
                stats = {'P': 0, 'A': 0, 'L': 0, 'T': 0, 'S': 0, '': 0, 'W': 0, 'H': 0}

                for date_obj in date_list:
                    date_str = date_obj.isoformat()
                    is_non_work, reason = is_non_working_day(date_obj, self.db)

                    status = att_data.get(roll, {}).get(date_str, '')

                    if is_non_work:
                        # 비근무일은 W/H로 카운트
                        status_key = 'H' if date_obj.weekday() < 5 else 'W'
                    else:
                        # 근무일은 실제 출결 상태 코드로 카운트
                        status_key = status

                    stats[status_key] += 1

                # 월간 누계 통계를 최종 데이터 리스트에 추가
                row_data = [roll, name, total_work_days] + [stats[k] for k in ['P', 'A', 'L', 'T', 'S', '', 'W', 'H']]
                data.append(row_data)

            # ----------------------------------------------------
            # PDF 테이블 출력
            # ----------------------------------------------------
            col_widths = [50, 100, 40] + [55] * 8
            table_style = TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), KOREAN_FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, 0), 1, colors.black),
                ('GRID', (0, 1), (-1, -1), 0.5, colors.gray),  # 데이터 행은 얇은 선으로 변경
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DDDDDD')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])

            table = Table(data, colWidths=col_widths)
            table.setStyle(table_style)
            Story.append(table)
            Story.append(Spacer(1, 12))

            legend_text = "출결 상태 범례: " + ", ".join([f"{k}: {v}" for k, v in STATUS_CODES.items() if k not in ('W', 'H')]) + ", W: 주말, H: 공휴일"
            Story.append(Paragraph(legend_text, styles['Normal']))

            doc.build(Story)

            messagebox.showinfo("PDF 출력 성공", f"월간 출석 누계 보고서가 '{file_path}'에 성공적으로 저장되었습니다.")
            self.set_status(f"{year}년 {month}월 월간 PDF 누계 보고서 출력 완료: {file_path}")

        except Exception as e:
            messagebox.showerror("오류", f"월간 PDF 보고서 생성 중 오류 발생: {e}")
            self.set_status(f"PDF 오류: {e}")

    # ==========================================================
    # 공휴일 관리
    # ==========================================================

    def add_holiday_dialog(self):
        """공휴일 추가 대화상자"""
        date_str = simpledialog.askstring("공휴일 추가",
                                          "날짜를 입력하세요 (YYYY-MM-DD):", parent=self.root)
        if not date_str:
            return

        try:
            datetime.date.fromisoformat(date_str)
        except:
            messagebox.showerror("오류", "올바른 날짜 형식이 아닙니다 (YYYY-MM-DD)")
            return

        name = simpledialog.askstring("공휴일 추가",
                                      "공휴일 이름을 입력하세요:", parent=self.root)
        if not name:
            return

        if self.db.add_holiday(date_str, name):
            messagebox.showinfo("성공", f"{date_str} - {name} 공휴일이 추가되었습니다.")
            self.set_status(f"공휴일 추가: {date_str} - {name}")
            self.load_attendance_for_date()
        else:
            messagebox.showerror("오류", "이미 등록된 공휴일입니다.")

    def import_holidays_from_excel(self):
        """Excel 파일에서 공휴일 불러와 DB에 저장 (월-일 형식 지원)"""
        if 'openpyxl' not in globals():
            messagebox.showerror("오류",
                                 "Excel 기능을 사용하려면 'openpyxl' 라이브러리가 필요합니다. \n"
                                 "명령 프롬프트에서 'pip install openpyxl'을 실행해주세요."
                                 )
            return

        try:
            file_path = filedialog.askopenfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="공휴일 데이터가 있는 Excel 파일을 선택하세요"
            )
            if not file_path:
                return

            if not messagebox.askyesno("확인",
                                       "기존에 등록된 공휴일은 유지되고, 엑셀 파일의 공휴일이 추가됩니다.\n"
                                       "계속하시겠습니까? (A열: 월-일, B열: 공휴일명)"
                                       ):
                return

            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            today_year = datetime.date.today().year
            years_to_check = [today_year, today_year + 1]

            added_count = 0
            skipped_count = 0

            for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                month_day_cell = row[0].value
                name_cell = row[1].value

                if not month_day_cell or not name_cell:
                    continue

                if isinstance(month_day_cell, datetime.datetime):
                    md_str = month_day_cell.strftime("%m-%d")
                else:
                    md_str = str(month_day_cell).strip()

                holiday_name = str(name_cell).strip()

                try:
                    parts = md_str.replace('/', '-').split('-')
                    month = int(parts[0])
                    day = int(parts[1])
                except (ValueError, IndexError):
                    self.set_status(f"⚠️ 경고: {row_index}행의 날짜 형식({md_str})이 올바르지 않아 건너뜀.")
                    skipped_count += 1
                    continue

                for year in years_to_check:
                    try:
                        full_date = datetime.date(year, month, day)
                        date_str = full_date.isoformat()

                        if self.db.add_holiday(date_str, holiday_name):
                            added_count += 1
                        else:
                            skipped_count += 1

                    except ValueError:
                        skipped_count += 1
                        continue

            messagebox.showinfo("Excel 불러오기 완료",
                                f"공휴일 불러오기 완료\n"
                                f"총 추가된 공휴일: {added_count}개 (중복 제외)\n"
                                f"건너뛴 항목: {skipped_count}개"
                                )
            self.set_status(f"Excel 공휴일 {added_count}개 추가됨")
            self.load_attendance_for_date()

        except Exception as e:
            messagebox.showerror("오류", f"Excel 파일 처리 중 오류 발생: {e}")
            self.set_status(f"Excel 오류: {e}")

    def view_holidays_dialog(self):
        """공휴일 목록 보기 및 관리 - 수정/추가/Excel불러오기 기능 통합"""
        holiday_win = Toplevel(self.root)
        holiday_win.title("공휴일 목록 관리")
        holiday_win.geometry("700x450")

        top_frame = ttk.Frame(holiday_win, padding=10)
        top_frame.pack(fill=X)
        ttk.Label(top_frame, text="등록된 공휴일 목록 (수정/삭제 가능)",
                  font=("맑은 고딕", 14, "bold")).pack()

        tree_frame = ttk.Frame(holiday_win)
        tree_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        holiday_tree = ttk.Treeview(tree_frame, columns=("date", "name"),
                                    show="headings", height=10)
        holiday_tree.heading("date", text="날짜")
        holiday_tree.column("date", width=150, anchor=CENTER)
        holiday_tree.heading("name", text="공휴일 이름")
        holiday_tree.column("name", width=450, anchor=W)

        scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL,
                                  command=holiday_tree.yview)
        holiday_tree.configure(yscroll=scrollbar.set)

        holiday_tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)

        def load_holidays():
            for item in holiday_tree.get_children():
                holiday_tree.delete(item)
            holidays = self.db.get_all_holidays()
            for date, name in holidays:
                holiday_tree.insert("", "end", values=(date, name))

        load_holidays()

        btn_frame = ttk.Frame(holiday_win, padding=10)
        btn_frame.pack(fill=X)

        def delete_selected():
            sel = holiday_tree.selection()
            if not sel:
                messagebox.showinfo("선택", "삭제할 공휴일을 선택하세요.")
                return
            date = holiday_tree.item(sel[0])["values"][0]
            if messagebox.askyesno("확인", f"{date} 공휴일을 삭제하시겠습니까?"):
                self.db.delete_holiday(date)
                load_holidays()
                self.set_status(f"공휴일 삭제: {date}")
                self.load_attendance_for_date()

        def edit_selected():
            sel = holiday_tree.selection()
            if not sel:
                messagebox.showinfo("선택", "수정할 공휴일을 선택하세요.")
                return

            item_data = holiday_tree.item(sel[0])["values"]
            date = item_data[0]
            current_name = item_data[1]

            new_name = simpledialog.askstring("공휴일 이름 수정",
                                              f"{date}의 새 공휴일 이름을 입력하세요:",
                                              initialvalue=current_name,
                                              parent=holiday_win)
            if new_name is None or new_name.strip() == "":
                return

            self.db.delete_holiday(date)

            if self.db.add_holiday(date, new_name.strip()):
                messagebox.showinfo("성공", f"{date}의 이름이 '{new_name}'으로 수정되었습니다.")
                load_holidays()
                self.set_status(f"공휴일 수정: {date} → {new_name}")
                self.load_attendance_for_date()
            else:
                self.db.add_holiday(date, current_name)
                messagebox.showerror("오류", "공휴일 수정 중 오류가 발생했습니다. (날짜 충돌)")
                load_holidays()

        def run_and_refresh(func):
            # Toplevel 윈도우에서 dialog를 띄우고 결과를 반영하기 위함
            func()
            load_holidays()

        ttk.Button(btn_frame, text="✅ 공휴일 추가",
                   command=lambda: run_and_refresh(self.add_holiday_dialog),
                   style="Success.TButton").pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="✏️ 공휴일 수정", command=edit_selected,
                   style="Accent.TButton").pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑️ 공휴일 삭제", command=delete_selected,
                   style="Danger.TButton").pack(side=LEFT, padx=5)

        ttk.Separator(btn_frame, orient=VERTICAL).pack(side=LEFT, padx=10, fill=Y)

        ttk.Button(btn_frame, text="📥 엑셀로 불러오기",
                   command=lambda: run_and_refresh(self.import_holidays_from_excel),
                   style="Accent.TButton").pack(side=LEFT, padx=5)

        ttk.Button(btn_frame, text="닫기", command=holiday_win.destroy).pack(
            side=RIGHT, padx=5)

        holiday_tree.bind("<Double-1>", lambda e: edit_selected())


# --- 프로그램 실행 ---
if __name__ == "__main__":
    # openpyxl 라이브러리 존재 여부 확인
    try:
        import openpyxl
    except ImportError:
        openpyxl = None

    root = Tk()
    app = AttendanceApp(root)

    # 윈도우 닫기 이벤트 핸들러
    root.protocol("WM_DELETE_WINDOW", app.on_close)

    root.mainloop()